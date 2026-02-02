import pandas as pd
from itertools import combinations
from collections import deque
from datetime import datetime
import warnings
import numpy as np

warnings.filterwarnings('ignore')
import sys
from openpyxl.styles import PatternFill, Alignment, Font # Needed for coloring and formatting
from openpyxl.chart import BarChart, Reference # Needed for charts

# --- CHANGE: Management of Numba as an optional dependency ---
try:
    from numba import jit
    NUMBA_AVAILABLE = True
except ImportError:
    NUMBA_AVAILABLE = False
    # Define a dummy decorator if Numba is not available
    def jit(signature_or_function=None, locals={}, cache=False, pipeline_class=None, boundscheck=None, **options):
        def decorator(func):
            return func
        return decorator

def _robust_currency_parser(value):
    """
    Robustly converts a string or number into a standard numeric format for pd.to_numeric.
    This helper function is used by `load_file`.
    """
    # If it's already a number, it's fine.
    if isinstance(value, (int, float)):
        return value
    # If it's not a string, we can't do anything.
    if not isinstance(value, str):
        return None # Will be converted to NaN
    
    # Clean the string from spaces and euro symbol
    cleaned_str = str(value).strip().replace('€', '').replace(' ', '')
    
    # Case 1: Full Italian format (e.g., "1.234,56")
    if '.' in cleaned_str and ',' in cleaned_str:
        return cleaned_str.replace('.', '').replace(',', '.')
    # Case 2: Italian format with only decimals (e.g., "1234,56")
    elif ',' in cleaned_str:
        return cleaned_str.replace(',', '.')
    # Case 3: Format without commas (e.g., "1234" or "1234.56"). We leave the dot.
    return cleaned_str

class RiconciliatoreContabile:
    """Contains the business logic for reconciliation."""

    def __init__(self, tolleranza=0.01, giorni_finestra=7, max_combinazioni=10, soglia_residui=100.0, giorni_finestra_residui=30, sorting_strategy="date", search_direction="past_only", column_mapping=None, algorithm="subset_sum", use_numba=True, ignore_tolerance=False, enable_best_fit=True):
        """
        Initializes the reconciler with configuration parameters.

        Args:
            tolleranza (float): Maximum accepted difference between amounts (default 0.01).
            giorni_finestra (int): Search time window in days (default 7).
            max_combinazioni (int): Maximum number of combinable movements (default 10).
            soglia_residui (float): Minimum amount to consider a movement in the residuals phase (default 100.0).
            giorni_finestra_residui (int): Extended time window for the residuals phase (default 30).
            sorting_strategy (str): Sorting strategy ('date' or 'amount').
            search_direction (str): Search direction ('past_only', 'future_only', 'both').
            column_mapping (dict): Column name mapping (e.g., {'Data': 'MyDate', ...}).
            algorithm (str): Algorithm to use ('subset_sum', 'progressive_balance', 'all').
            use_numba (bool): If True, uses Numba acceleration if available.
            ignore_tolerance (bool): If True, forces closing blocks in progressive balance even if they don't match.
            enable_best_fit (bool): If True, enables partial matching logic (splitting).
        """
        # FIX: Converts values from euros (float) to cents (int) for internal consistency
        self.tolleranza = int(tolleranza * 100)
        self.giorni_finestra = giorni_finestra
        self.max_combinazioni = max_combinazioni
        # FIX: Converts values from euros (float) to cents (int)
        self.soglia_residui = int(soglia_residui * 100)
        self.giorni_finestra_residui = giorni_finestra_residui
        self.sorting_strategy = sorting_strategy
        self.search_direction = search_direction
        self.algorithm = algorithm
        # ADDITION: Sets the column mapping, with a default if not provided.
        self.column_mapping = column_mapping or {'Data': 'Data', 'Dare': 'Dare', 'Avere': 'Avere'}
        
        # Flag to enable/disable Numba
        self.use_numba = use_numba and NUMBA_AVAILABLE
        
        # Flag to force closing blocks in Progressive Balance even if they don't match (on window timeout)
        self.ignore_tolerance = ignore_tolerance

        # ADDITION: Flag to enable best-fit logic (splitting)
        self.enable_best_fit = enable_best_fit

        # Internal state that will be populated during execution
        self.dare_df = self.avere_df = self.df_abbinamenti = None
        self.dare_non_util = self.avere_non_riconc = self.original_df = None

        # Optimization: Use sets to keep track of used indices
        self.used_dare_indices = set()
        self.used_avere_indices = set()
        
        # Counter to generate new unique IDs for residuals
        self.max_id_counter = 0

    def carica_file(self, file_path):
        """
        Loads an Excel, CSV, or Feather file into a standardized Pandas DataFrame.

        Handles reading, renaming columns according to the configured mapping,
        date conversion, and amount cleaning.

        Args:
            file_path (str): Path of the file to load.

        Returns:
            pd.DataFrame: DataFrame with standardized columns ('Data', 'Dare', 'Avere', 'indice_orig').
        """
        # Common parameters for reading CSV/Excel with European format
        common_read_params = {
            'decimal': ',',
            'thousands': '.'
        }

        if str(file_path).endswith('.csv'):
            df = pd.read_csv(file_path, decimal=',', thousands='.')
        elif str(file_path).endswith('.feather'):
            df = pd.read_feather(file_path)
        else:
            # This branch should only be reached if it's not a feather and not a CSV.
            # For Excel files, convert_to_feather.py should have already handled the parsing.
            df = pd.read_excel(file_path, decimal=',', thousands='.', engine='openpyxl')

        # --- CHANGE: Dynamic handling of column names ---
        # Invert the map for renaming: {'Source Column Name': 'Internal Name'} -> {'Internal Name': 'Source Column Name'}
        source_col_names = self.column_mapping.keys()
        
        # Check if the source columns defined in the configuration exist in the file
        if not set(source_col_names).issubset(df.columns):
            missing_cols = set(source_col_names) - set(df.columns)
            raise ValueError(f"The input file does not contain the source columns specified in the configuration: {', '.join(missing_cols)}")

        # Rename the DataFrame columns using the mapping to standardize them to internal names ('Data', 'Dare', 'Avere')
        df.rename(columns=self.column_mapping, inplace=True)

        # After reading, ensure 'Data' is datetime and 'Dare'/'Avere' are numeric.
        # This is a fallback in case the reading parameters were not sufficient
        # or if the DataFrame comes from an already pre-loaded source (e.g., from the optimizer).
        df['Data'] = pd.to_datetime(df['Data'], errors='coerce', dayfirst=True)
        df.dropna(subset=['Data'], inplace=True) # Removes rows with invalid dates

        # --- CHECK DATE FUTURE ---
        today = datetime.now()
        future_rows = df[df['Data'] > today]
        if not future_rows.empty:
            print(f"\n⚠️  WARNING: Found {len(future_rows)} movements with a future date (compared to {today.strftime('%d/%m/%Y')})!")
            print(f"    Example: {future_rows.iloc['Data'].strftime('%d/%m/%Y')} (Row {future_rows.index + 2})")

        # --- ROBUST AMOUNT CLEANING ---
        # Apply the robust parser to each cell, then convert the entire column.
        for col in ['Dare', 'Avere']:
            df[col] = pd.to_numeric(df[col].apply(_robust_currency_parser), errors='coerce')

        # Fill non-numeric values with 0 BEFORE converting to cents
        df[['Dare', 'Avere']] = df[['Dare', 'Avere']].fillna(0)

        # --- OPTIMIZATION: Convert to integers (cents) to avoid floating point errors ---
        # We multiply by 100 and round for safety, then convert to integers.
        df['Dare'] = (df['Dare'] * 100).round().astype(int)
        df['Avere'] = (df['Avere'] * 100).round().astype(int)

        df['indice_orig'] = df.index
        
        # Initialize the ID counter with the maximum existing index
        if not df.empty:
            self.max_id_counter = df.index.max()
            
        return df

    def _separa_movimenti(self, df):
        """Separates the DataFrame into DEBIT and CREDIT movements."""
        self.dare_df = df[df['Dare'] != 0][['indice_orig', 'Data', 'Dare']].copy()
        self.avere_df = df[df['Avere'] != 0][['indice_orig', 'Data', 'Avere']].copy()
       
        if self.sorting_strategy == "date":
            self.dare_df = self.dare_df.sort_values('Data', ascending=True)
            self.avere_df = self.avere_df.sort_values('Data', ascending=True)
        elif self.sorting_strategy == "amount":
            self.dare_df = self.dare_df.sort_values('Dare', ascending=False)
            self.avere_df = self.avere_df.sort_values('Avere', ascending=False)
        else:
            raise ValueError(f"Invalid sorting strategy: '{self.sorting_strategy}'. Use 'date' or 'amount'.")
            
        return self.dare_df, self.avere_df

    def _trova_abbinamenti(self, dare_row, avere_candidati_np, avere_indices_map, giorni_finestra, max_combinazioni, enable_best_fit=False):
        """Internal logic to find a match for a single DEBIT. Receives pre-filtered candidates by date."""
        dare_importo = dare_row['Dare']
        dare_data = dare_row['Data']

        # --- CORRECT OPTIMIZATION: Filter the list of dictionaries ---
        # avere_candidati_np is now a list of dictionaries, not a numpy array
        # Candidates are already pre-filtered by date and unused indices. We only filter by amount.
        candidati_avere = [
            c for c in avere_candidati_np if c['Avere'] <= dare_importo + self.tolleranza
        ]
        
        if not candidati_avere:
            return None
        
        # 1. Search for exact 1-to-1 match
        match_esatto_list = [c for c in candidati_avere if abs(c['Avere'] - dare_importo) <= self.tolleranza]
        if match_esatto_list:
            best_match = match_esatto_list[0] # Prende il primo match esatto trovato
            return {
                'dare_indices': [dare_row['indice_orig']],
                'dare_date': [dare_data],
                'dare_importi': [dare_importo],
                'avere_indices': [best_match['indice_orig']],
                'avere_date': [best_match['Data']],
                'avere_importi': [best_match['Avere']],
                'somma_avere': best_match['Avere'],
                'differenza': abs(dare_importo - best_match['Avere']),
                'tipo_match': '1-a-1'
            }

        # 2. Search for multiple combinations in an optimized way
        candidati_avere = sorted(candidati_avere, key=lambda x: x['Avere'], reverse=True)
        
        # Added cache for memoization
        cache = {}
        somma_totale_candidati = sum(c['Avere'] for c in candidati_avere)
        
        match = None
        if self.use_numba:
            # --- NUMBA OPTIMIZATION ---
            candidati_np_numba = np.array([(c['Avere'], c['indice_orig']) for c in candidati_avere], dtype=np.int64)
            match_indices = _numba_find_combination(dare_importo, candidati_np_numba, max_combinazioni, self.tolleranza)
            if len(match_indices) > 0:
                match = [c for c in candidati_avere if c['indice_orig'] in match_indices]
        else:
            # --- FALLBACK TO PURE PYTHON ---
            for c in candidati_avere: c['Dare'] = c.pop('Avere') # Adatta per la funzione generica
            match = self._trova_combinazioni_ricorsivo_py(dare_importo, candidati_avere, max_combinazioni, self.tolleranza)
            if match:
                for c in match: c['Avere'] = c.pop('Dare') # Ripristina

        if match:
            return {
                'dare_indices': [dare_row['indice_orig']],
                'dare_date': [dare_data],
                'dare_importi': [dare_importo],
                'avere_indices': [m['indice_orig'] for m in match],
                'avere_date': [m['Data'] for m in match],
                'avere_importi': [m['Avere'] for m in match],
                'somma_avere': sum(m['Avere'] for m in match),
                'differenza': abs(dare_importo - sum(m['Avere'] for m in match)),
                'tipo_match': f'Combination {len(match)}'
            }
        
        return None

    def _trova_combinazioni_ricorsivo_py(self, target, candidati, max_combinazioni, tolleranza):
        """Iterative function (stack-based) for subset-sum. Pure Python version."""
        stack = deque([(0, 0, [])]) # (start_index, somma_parziale, percorso_parziale)

        while stack:
            idx, current_sum, current_path = stack.pop()

            # --- SUCCESS CONDITION ---
            if abs(target - current_sum) <= tolleranza and len(current_path) > 1:
                return current_path

            # --- PRUNING CONDITIONS ---
            if len(current_path) >= max_combinazioni or idx >= len(candidati):
                continue

            # --- BRANCH 1: Exclude the current candidate ---
            # Continue exploration from the next candidate.
            stack.append((idx + 1, current_sum, current_path))

            # --- BRANCH 2: Include the current candidate ---
            candidato = candidati[idx]
            new_sum = current_sum + candidato['Dare'] # The generic function uses 'Dare'
            
            # Pruning: do not include if the new sum already exceeds too much
            if new_sum > target + tolleranza:
                continue

            new_path = current_path + [candidato]
            
            # --- SUCCESS CONDITION (even after addition) ---
            if abs(target - new_sum) <= tolleranza and len(new_path) > 1:
                return new_path

            # Continue exploration including the current element
            stack.append((idx + 1, new_sum, new_path))

        return None # No combination found

    def _trova_abbinamenti_dare(self, avere_row, dare_candidati_np, dare_indices_map, giorni_finestra, max_combinazioni, enable_best_fit=False):
        """Logic to find combinations of DEBIT that match a CREDIT. Receives pre-filtered candidates."""
        avere_importo = avere_row['Avere']
        avere_data = avere_row['Data']

        # Candidates are already pre-filtered by date and unused indices. We only filter by amount.
        candidati_dare_list = [c for c in dare_candidati_np if c['Dare'] <= avere_importo + self.tolleranza]

        if not candidati_dare_list:
            return None

        # Search for multiple DEBIT combinations in an optimized way
        candidati_da_modificare = [c.copy() for c in candidati_dare_list]
        candidati_da_modificare = sorted(candidati_da_modificare, key=lambda x: x['Dare'], reverse=True)

        match = None
        is_partial = False
        
        if self.use_numba:
            candidati_np_numba = np.array([(c['Dare'], c['indice_orig']) for c in candidati_da_modificare], dtype=np.int64)
            # Attempt 1: Exact Match
            match_indices = _numba_find_combination(avere_importo, candidati_np_numba, max_combinazioni, self.tolleranza)
            
            if len(match_indices) > 0:
                match = [c for c in candidati_da_modificare if c['indice_orig'] in match_indices]
            elif enable_best_fit:
                # Attempt 2: Best Fit (Partial Match)
                # Find the combination that best fills the payment without exceeding it
                match_indices = _numba_find_best_fit_combination(avere_importo, candidati_np_numba, max_combinazioni, self.tolleranza)
                if len(match_indices) > 0:
                    match = [c for c in candidati_da_modificare if c['indice_orig'] in match_indices]
                    is_partial = True
        else:
            match = self._trova_combinazioni_ricorsivo_py(avere_importo, candidati_da_modificare, max_combinazioni, self.tolleranza)

        if match:
            somma_dare = sum(m['Dare'] for m in match)
            differenza = abs(avere_importo - somma_dare)
            
            # If it's a partial best fit, we calculate the residual
            residuo = 0
            if is_partial and differenza > self.tolleranza:
                residuo = differenza
            
            return {
                'dare_indices': [m['indice_orig'] for m in match],
                'dare_date': [m['Data'] for m in match],
                'dare_importi': [m['Dare'] for m in match],
                'avere_indices': [avere_row['indice_orig']],
                'avere_date': [avere_data],
                'avere_importi': [avere_importo],
                'somma_dare': somma_dare,
                'differenza': differenza,
                'tipo_match': f'DEBIT Combination {len(match)}' + (' (Best Fit)' if is_partial else ''),
                'residuo': residuo if is_partial else 0
            }
        return None

    def _esegui_passata_riconciliazione_dare(self, dare_df, avere_df, giorni_finestra, max_combinazioni, abbinamenti_list, title, verbose=True, enable_best_fit=False):
        """Runs a pass looking for DEBIT combinations to match CREDIT (optimized with NumPy)."""
        # Filter CREDITs that have not been used yet
        avere_da_processare = avere_df[~avere_df['indice_orig'].isin(self.used_avere_indices)].copy() if avere_df is not None and not avere_df.empty else pd.DataFrame()

        self._esegui_passata_generica(
            df_da_processare=avere_da_processare,
            df_candidati=dare_df,
            col_da_processare='Avere',
            col_candidati='Dare',
            used_indices_candidati=self.used_dare_indices,
            giorni_finestra=giorni_finestra,
            max_combinazioni=max_combinazioni,
            abbinamenti_list=abbinamenti_list,
            title=title,
            search_direction=self.search_direction, # Use the main direction from the configuration
            find_function=self._trova_abbinamenti_dare,
            verbose=verbose,
            enable_best_fit=enable_best_fit
        )

    def _esegui_passata_generica(self, df_da_processare, df_candidati, col_da_processare, col_candidati, used_indices_candidati, giorni_finestra, max_combinazioni, abbinamenti_list, title, search_direction, find_function, verbose=True, enable_best_fit=False):
        """
        Generic helper function that performs a reconciliation pass.
        
        Iterates over each row of `df_da_processare` and searches for matches in `df_candidati`
        using the provided `find_function`. Manages time window logic,
        match registration, and splitting (Best Fit).

        Args:
            df_da_processare (pd.DataFrame): Main DataFrame to iterate over.
            df_candidati (pd.DataFrame): DataFrame where to search for combinations.
            col_da_processare (str): Amount column name in the main DF ('Dare' or 'Avere').
            col_candidati (str): Amount column name in the candidate DF.
            used_indices_candidati (set): Set of already used indices to exclude.
            giorni_finestra (int): Time window for the search.
            max_combinazioni (int): Max combinable elements.
            abbinamenti_list (list): List to append found matches to.
            title (str): Title of the pass for logging.
            search_direction (str): Time direction ('past_only', 'future_only', 'both').
            find_function (callable): Function that implements the specific matching logic.
            verbose (bool): If True, prints logs.
            enable_best_fit (bool): If True, enables splitting logic for partial matches.
        """
        if df_da_processare is None or df_da_processare.empty:
            return

        if verbose:
            print(f"\n{title} (Direction: {search_direction})...")

        # Prepare the record lists once
        records_da_processare = df_da_processare.to_dict('records')
        records_candidati = sorted(df_candidati.to_dict('records'), key=lambda x: x['Data']) if df_candidati is not None else []

        matches = []
        total_records = len(records_da_processare)
        processed_count = 0
        
        # List to collect new residual movements generated by splitting
        nuovi_residui = []

        for record_row in records_da_processare:
            if verbose:
                processed_count += 1
                percentuale = (processed_count / total_records) * 100
                sys.stdout.write(f"\r   - Progress: {percentuale:.1f}% ({processed_count}/{total_records})")
                sys.stdout.flush()

            # Pre-filter candidates by time window
            min_data, max_data = self._calcola_finestra_temporale(record_row['Data'], giorni_finestra, search_direction)
            
            candidati_prefiltrati = [
                c for c in records_candidati
                if min_data <= c['Data'] <= max_data and c['indice_orig'] not in used_indices_candidati
            ]

            if candidati_prefiltrati:
                match = find_function(record_row, candidati_prefiltrati, None, giorni_finestra, max_combinazioni, enable_best_fit=enable_best_fit)
                if match:
                    # Handle split (Best Fit)
                    residuo = match.get('residuo', 0)
                    if residuo > 0:
                        # Create a new movement for the residual
                        nuovo_movimento = self._crea_movimento_residuo(record_row, residuo, col_da_processare)
                        nuovi_residui.append(nuovo_movimento)

                        # Update the amount of the original row to reflect only the reconciled part
                        # This corrects the statistics by avoiding amount duplication (Original + Residual)
                        idx_orig = record_row['indice_orig']
                        new_amount = record_row[col_da_processare] - residuo
                        
                        if col_da_processare == 'Avere':
                            self.avere_df.loc[self.avere_df['indice_orig'] == idx_orig, 'Avere'] = new_amount
                            # FIX REPORT: Also update the match to show only the used part in Excel
                            match['avere_importi'] = [new_amount]
                            match['somma_avere'] = new_amount
                            match['differenza'] = abs(match.get('somma_dare', 0) - new_amount)
                        elif col_da_processare == 'Dare':
                            self.dare_df.loc[self.dare_df['indice_orig'] == idx_orig, 'Dare'] = new_amount
                            # FIX REPORT
                            match['dare_importi'] = [new_amount]
                            match['somma_dare'] = new_amount
                            match['differenza'] = abs(match.get('somma_avere', 0) - new_amount)

                    # --- CRITICAL FIX: IMMEDIATE REGISTRATION ---
                    # Immediately register the match to mark indices as used and prevent
                    # them from being reused in the same pass (Double Spending).
                    match['pass_name'] = title
                    self._registra_abbinamento(match, abbinamenti_list)
                    matches.append(match) # Keeps the list only for the final count

        if verbose: print(f"\n   - Registered {len(matches)} matches.")
            
        # Add the generated residuals to the original DataFrame to be processed in subsequent passes
        if nuovi_residui:
            if verbose: print(f"   - Generated {len(nuovi_residui)} residual movements from split (Best Fit).")
            df_residui = pd.DataFrame(nuovi_residui)
            
            if col_da_processare == 'Avere':
                self.avere_df = pd.concat([self.avere_df, df_residui], ignore_index=True)
                # Ensure types are correct
                self.avere_df['Avere'] = self.avere_df['Avere'].astype(int)
            elif col_da_processare == 'Dare':
                self.dare_df = pd.concat([self.dare_df, df_residui], ignore_index=True)
                self.dare_df['Dare'] = self.dare_df['Dare'].astype(int)
                
        if verbose: sys.stdout.write("\n   ✓ Completed.\n")

    def _crea_movimento_residuo(self, record_originale, importo_residuo, col_tipo):
        """Creates a dictionary representing the residual movement."""
        self.max_id_counter += 1
        nuovo_id = self.max_id_counter
        
        nuovo_movimento = record_originale.copy()
        nuovo_movimento['indice_orig'] = nuovo_id
        nuovo_movimento[col_tipo] = importo_residuo
        # Note: 'usato' will be False (or NaN which will be treated as False) by default when added to the DF
        
        return nuovo_movimento

    def _calcola_finestra_temporale(self, data_riferimento, giorni_finestra, search_direction):
        """Calculates the time window (min_data, max_data) based on the search direction."""
        if search_direction == "future_only":
            min_data = data_riferimento
            max_data = data_riferimento + pd.Timedelta(days=giorni_finestra)
        elif search_direction == "past_only":
            min_data = data_riferimento - pd.Timedelta(days=giorni_finestra)
            max_data = data_riferimento
        elif search_direction == "both":
            min_data = data_riferimento - pd.Timedelta(days=giorni_finestra)
            max_data = data_riferimento + pd.Timedelta(days=giorni_finestra)
        else:
            raise ValueError(f"Invalid time search direction: '{search_direction}'. Use 'both', 'future_only' or 'past_only'.")
        return min_data, max_data

    def _esegui_passata_riconciliazione(self, dare_df, avere_df, giorni_finestra, max_combinazioni, abbinamenti_list, title, verbose=True):
        """Performs a reconciliation pass and updates the DataFrames (optimized with NumPy)."""
        # Filter DEBITs that have not been used yet
        dare_da_processare = dare_df[~dare_df['indice_orig'].isin(self.used_dare_indices)].copy() if dare_df is not None and not dare_df.empty else pd.DataFrame()
        
        # --- FIX: Logical inversion of the direction for the DEBIT->CREDIT pass ---
        # If the global strategy is "past_only" (DEBIT before CREDIT),
        # when we start from DEBIT we must search for CREDIT in the future ("future_only").
        direction_for_pass2 = self.search_direction
        if self.search_direction == "past_only":
            direction_for_pass2 = "future_only"
        elif self.search_direction == "future_only":
            direction_for_pass2 = "past_only"
        
        self._esegui_passata_generica(
            df_da_processare=dare_da_processare,
            df_candidati=avere_df,
            col_da_processare='Dare',
            col_candidati='Avere',
            used_indices_candidati=self.used_avere_indices,
            giorni_finestra=giorni_finestra,
            max_combinazioni=max_combinazioni,
            abbinamenti_list=abbinamenti_list,
            title=title,
            search_direction=direction_for_pass2, # Use the correct (inverted) direction
            find_function=self._trova_abbinamenti,
            verbose=verbose
        )

    def _riconcilia_subset_sum(self, verbose=True):
        """
        Esegue la riconciliazione basata sulla ricerca di combinazioni (Subset Sum).
        Performs reconciliation based on combination search (Subset Sum).

        Orchestrates three successive passes:
        1. Many DEBIT -> 1 CREDIT (with optional Best Fit).
        2. 1 DEBIT -> Many CREDIT (Split payments).
        3. Residual recovery with an extended time window.

        Args:
            verbose (bool): If True, prints progress.

        Returns:
            list: List of dictionaries representing the found matches.
        """
        
        abbinamenti = []

        # --- Pass 1: DEBIT combination for CREDIT (Many Receipts -> 1 Deposit) ---
        # This is the "Human" logic: I look for which past receipts make up this deposit.
        # Highest priority after implicit 1-to-1 matches.
        # BEST FIT ENABLED: If it doesn't find an exact match, it tries to partially fill the deposit.
        self._esegui_passata_riconciliazione_dare(
            self.dare_df, self.avere_df,
            self.giorni_finestra,
            self.max_combinazioni,
            abbinamenti,
            "Pass 1: Receipt Aggregation (Many DEBIT -> 1 CREDIT) [with Best Fit]",
            verbose,
            enable_best_fit=self.enable_best_fit
        )

        # --- Pass 2: Standard Inverse Reconciliation (1 Receipt -> Many Deposits) ---
        # Useful if a very large receipt is deposited in several installments (less common but possible).
        self._esegui_passata_riconciliazione(
            self.dare_df, self.avere_df,
            self.giorni_finestra,
            self.max_combinazioni,
            abbinamenti,
            "Pass 2: Split Deposits (1 DEBIT -> Many CREDIT)",
            verbose
        )

        # --- Pass 3: Residual Analysis (Enlarged Window) ---
        # Tries to recover what was left out with a wider window
        self._esegui_passata_riconciliazione_dare(
            self.dare_df, self.avere_df,
            self.giorni_finestra_residui,
            self.max_combinazioni,
            abbinamenti,
            f"Pass 3: Residual Recovery (Extended window: {self.giorni_finestra_residui}d)",
            verbose,
            enable_best_fit=False
        )

        return abbinamenti

    def _finalizza_abbinamenti(self, abbinamenti):
        """Creates the final DataFrame, generates IDs, and sorts the results."""
        # Expected columns in the final DataFrame
        final_columns = [
            'ID Transazione', 'dare_indices', 'dare_date', 'dare_importi', 
            'avere_data', 'num_avere', 'avere_indices', 'avere_importi', 
            'somma_avere', 'differenza', 'giorni_diff', 'tipo_match', 'pass_name'
        ]

        # Creation of the final matches DataFrame
        if abbinamenti:
            df_abbinamenti = pd.DataFrame(abbinamenti)
            # Handling of missing columns (e.g., 'somma_dare' vs 'somma_avere')
            if 'somma_dare' in df_abbinamenti.columns and 'somma_avere' not in df_abbinamenti.columns:
                df_abbinamenti['somma_avere'] = df_abbinamenti['somma_dare']
            
            # Calculation of day difference (Credit - Debit)
            df_abbinamenti['giorni_diff'] = df_abbinamenti.apply(
                lambda row: (row['avere_data'] - min(row['dare_date'])).days 
                if isinstance(row['dare_date'], list) and len(row['dare_date']) > 0 and pd.notnull(row['avere_data']) 
                else None, axis=1
            )

            # --- CHANGE: Creation of the Transaction ID with new format D(..)_A(..) ---
            df_abbinamenti['ID Transazione'] = df_abbinamenti.apply(
                lambda row: "D({})_A({})".format(
                    ','.join(map(str, [i + 2 for i in row['dare_indices']])),
                    ','.join(map(str, [i + 2 for i in row['avere_indices']]))
                ), axis=1
            )
            df_abbinamenti['sort_date'] = df_abbinamenti['dare_date'].apply(lambda x: x[0] if isinstance(x, list) else x)
            df_abbinamenti['sort_importo'] = df_abbinamenti['dare_importi'].apply(lambda x: sum(x) if isinstance(x, list) else x)
            df_abbinamenti = df_abbinamenti.sort_values(by=['sort_date', 'sort_importo'], ascending=[True, False]).drop(columns=['sort_date', 'sort_importo'])
            df_abbinamenti = df_abbinamenti.reindex(columns=final_columns) # Ensures all columns exist
        else:
            df_abbinamenti = pd.DataFrame(columns=final_columns)
            
        return df_abbinamenti

    def _riconcilia_saldo_progressivo(self, verbose=True):
        """
        Reconciliation algorithm based on sequential progressive balance (Two Pointers).

        Simulates an operator who scrolls through chronologically ordered lists and closes
        a block when the progressive sum of DEBIT and CREDIT matches.

        Args:
            verbose (bool): If True, prints progress.

        Returns:
            list: List of dictionaries representing the found matches (blocks).
        """
        from datetime import timedelta # Make sure it's imported
        if verbose:
            print("\nStarting reconciliation with 'Progressive Balance' algorithm (Sequential)...")

        # 1. Prepare data: Filter unused and Sort by Date
        # Note: We use copies to not modify the original dfs during iteration
        df_dare_temp = self.dare_df[~self.dare_df['indice_orig'].isin(self.used_dare_indices)].copy()
        df_avere_temp = self.avere_df[~self.avere_df['indice_orig'].isin(self.used_avere_indices)].copy()
        
        # Sorting by date: it is fundamental and intentional for this algorithm.
        # The algorithm simulates a balance that progresses over time, so it ignores the global
        # 'sorting_strategy' (e.g., by amount) and always forces a chronological order.
        df_dare_temp.sort_values(by=['Data', 'indice_orig'], inplace=True)
        df_avere_temp.sort_values(by=['Data', 'indice_orig'], inplace=True)

        dare_rows = df_dare_temp.to_dict('records')
        avere_rows = df_avere_temp.to_dict('records')
        
        n_dare = len(dare_rows)
        n_avere = len(avere_rows)
        
        i = 0 # Debit pointer
        j = 0 # Credit pointer
        
        cum_dare = 0
        cum_avere = 0
        
        start_i = 0
        start_j = 0
        
        abbinamenti = []
        
        if verbose:
            print(f"   - Sequential analysis on {n_dare} Debit movements and {n_avere} Credit movements...")

        # Main loop (Two Pointers)
        while i < n_dare or j < n_avere:
            # Check if we have reached a break-even point (with at least one movement processed in the current block)
            diff = cum_dare - cum_avere
            
            # --- IMPROVED RESET LOGIC (Block Duration) ---
            # Calculate the duration of the block accumulated so far
            # Start date: minimum between the first DEBIT and the first CREDIT of the current block
            start_date_dare = dare_rows[start_i]['Data'] if start_i < n_dare else None
            start_date_avere = avere_rows[start_j]['Data'] if start_j < n_avere else None
            
            # End date: maximum between the last processed DEBIT and CREDIT (i-1, j-1) or current
            curr_date_dare = dare_rows[i]['Data'] if i < n_dare else (dare_rows[i-1]['Data'] if i > 0 else None)
            curr_date_avere = avere_rows[j]['Data'] if j < n_avere else (avere_rows[j-1]['Data'] if j > 0 else None)
            
            valid_starts = [d for d in [start_date_dare, start_date_avere] if d is not None]
            valid_ends = [d for d in [curr_date_dare, curr_date_avere] if d is not None]
            
            should_reset = False
            if valid_starts and valid_ends:
                block_duration = (max(valid_ends) - min(valid_starts)).days
                if block_duration > self.giorni_finestra and (cum_dare > 0 or cum_avere > 0):
                    should_reset = True

            if should_reset:
                if self.ignore_tolerance:
                    # --- FORCE CLOSE (Accept error) ---
                    # If the user has chosen to ignore the tolerance (or rather, to force on timeout),
                    # we close the block as is.
                    block_dare = dare_rows[start_i:i]
                    block_avere = avere_rows[start_j:j]
                    match = {
                        'dare_indices': [r['indice_orig'] for r in block_dare],
                        'dare_date': [r['Data'] for r in block_dare],
                        'dare_importi': [r['Dare'] for r in block_dare],
                        'avere_indices': [r['indice_orig'] for r in block_avere],
                        'avere_date': [r['Data'] for r in block_avere],
                        'avere_importi': [r['Avere'] for r in block_avere],
                        'somma_avere': cum_avere,
                        'differenza': abs(diff),
                        'tipo_match': f'Forced (Timeout {self.giorni_finestra}d)',
                        'pass_name': 'Progressive Balance (Forced)'
                    }
                    self._registra_abbinamento(match, abbinamenti)
                    # Reset and continue
                    start_i = i
                    start_j = j
                    cum_dare = 0
                    cum_avere = 0
                else:
                    # --- STANDARD RESET (Skip incorrect block) ---
                    # Abandon the current block that doesn't balance and start fresh.
                    # This allows finding subsequent matches instead of carrying over the error.
                    start_i = i
                    start_j = j
                    cum_dare = 0
                    cum_avere = 0
            
            # Match Condition: Zero difference (within tolerance) and we have advanced at least one of the pointers
            if abs(diff) <= self.tolleranza and (i > start_i or j > start_j):
                # --- BALANCED BLOCK FOUND ---
                block_dare = dare_rows[start_i:i]
                block_avere = avere_rows[start_j:j]
                
                match = {
                    'dare_indices': [r['indice_orig'] for r in block_dare],
                    'dare_date': [r['Data'] for r in block_dare],
                    'dare_importi': [r['Dare'] for r in block_dare],
                    'avere_indices': [r['indice_orig'] for r in block_avere],
                    'avere_date': [r['Data'] for r in block_avere],
                    'avere_importi': [r['Avere'] for r in block_avere],
                    'somma_avere': cum_avere, # O cum_dare, sono uguali
                    'differenza': abs(diff),
                    'tipo_match': f'Progressive Balance (Seq. {len(block_dare)}D vs {len(block_avere)}C)',
                    'pass_name': 'Progressive Balance'
                }
                self._registra_abbinamento(match, abbinamenti)

                # Reset for the next block (we start from zero to avoid error accumulation)
                start_i = i
                start_j = j
                cum_dare = 0
                cum_avere = 0
                
                # If we have finished both, we exit
                if i == n_dare and j == n_avere:
                    break
            
            # --- ADVANCEMENT LOGIC (GREEDY) ---
            # We decide which pointer to advance to try to balance the accounts.
            
            can_advance_dare = i < n_dare
            can_advance_avere = j < n_avere
            
            if can_advance_dare and can_advance_avere:
                # If Debit amount is behind, we add Debit
                if cum_dare < cum_avere:
                    cum_dare += dare_rows[i]['Dare']
                    i += 1
                # If Credit amount is behind, we add Credit
                elif cum_avere < cum_dare:
                    cum_avere += avere_rows[j]['Avere']
                    j += 1
                else:
                    # If amounts are equal (start of block or zero amounts), we advance the one with the earlier date
                    date_dare = dare_rows[i]['Data']
                    date_avere = avere_rows[j]['Data']
                    
                    if date_dare <= date_avere:
                        cum_dare += dare_rows[i]['Dare']
                        i += 1
                    else:
                        cum_avere += avere_rows[j]['Avere']
                        j += 1
                        
            elif can_advance_dare:
                # We can only advance Debit
                cum_dare += dare_rows[i]['Dare']
                i += 1
            elif can_advance_avere:
                # We can only advance Credit
                cum_avere += avere_rows[j]['Avere']
                j += 1
            else:
                # We can't advance either, but we haven't matched (case of non-squared final residual)
                break

        if verbose:
            print(f"   - Found {len(abbinamenti)} balanced blocks.")
            
        return abbinamenti

    def _registra_abbinamento(self, match, abbinamenti_list): # Rimosso dare_df, avere_df dagli argomenti
        """Marks the elements as 'used' and registers the match."""
        if not match:
            return

        dare_indices_orig = match.get('dare_indices', [])
        avere_indices_orig = match.get('avere_indices', [])

        # Add the indices to the sets of used indices
        self.used_dare_indices.update(dare_indices_orig)
        self.used_avere_indices.update(avere_indices_orig)

        # Add to formatted results
        # Ensure 'pass_name' is included
        avere_dates = match.get('avere_date')
        abbinamenti_list.append({
            'dare_indices': dare_indices_orig,
            'dare_date': match.get('dare_date', []),
            'dare_importi': match.get('dare_importi', []),
            'avere_data': min(avere_dates) if avere_dates else None,
            'num_avere': len(avere_indices_orig),
            'avere_indices': avere_indices_orig,
            'avere_importi': match.get('avere_importi', []),
            'somma_avere': match.get('somma_avere', match.get('somma_dare', 0)),
            'differenza': match.get('differenza', 0),
            'tipo_match': match.get('tipo_match', 'N/D'),
            'pass_name': match.get('pass_name', 'N/D')
        })

    def _calcola_quadratura_mensile(self):
        """Calculates aggregate statistics by month to identify periodic imbalances."""
        if self.dare_df is None or self.avere_df is None:
            return pd.DataFrame()

        # Helper to group
        def aggrega(df, col_valore):
            if df.empty:
                return pd.DataFrame()
            temp = df.copy()
            # Make sure Data is datetime
            temp['Data'] = pd.to_datetime(temp['Data'])
            temp['Mese'] = temp['Data'].dt.to_period('M')
            
            # Group by month
            gruppo = temp.groupby('Mese')
            
            totale = gruppo[col_valore].sum()
            usato = temp[temp['usato']].groupby('Mese')[col_valore].sum()
            
            res = pd.DataFrame({
                f'Totale {col_valore}': totale,
                f'Usato {col_valore}': usato
            })
            return res.fillna(0)

        stats_dare = aggrega(self.dare_df, 'Dare') # Columns: Totale Dare, Usato Dare
        stats_avere = aggrega(self.avere_df, 'Avere') # Columns: Totale Avere, Usato Avere

        # Union of the two dataframes (outer join to cover all months)
        stats = pd.merge(stats_dare, stats_avere, left_index=True, right_index=True, how='outer')
        
        # --- NEW: Calculation of the absorbed imbalance in matches ---
        sbilancio_assorbito = pd.DataFrame()
        if self.df_abbinamenti is not None and not self.df_abbinamenti.empty:
            df_temp_abbinamenti = self.df_abbinamenti.copy()
            
            # The reference date for the month is the date of the first DEBIT in the block
            df_temp_abbinamenti['Mese'] = df_temp_abbinamenti['dare_date'].apply(
                lambda x: x[0].to_period('M') if isinstance(x, list) and x else None
            )
            df_temp_abbinamenti.dropna(subset=['Mese'], inplace=True)
            
            # Calculate the signed difference (DEBIT - CREDIT) for each block
            df_temp_abbinamenti['somma_dare'] = df_temp_abbinamenti['dare_importi'].apply(lambda x: sum(x) if isinstance(x, list) else 0)
            df_temp_abbinamenti['sbilancio_blocco'] = df_temp_abbinamenti['somma_dare'] - df_temp_abbinamenti['somma_avere']
            
            sbilancio_assorbito = df_temp_abbinamenti.groupby('Mese')['sbilancio_blocco'].sum().to_frame('Absorbed Imbalance (in match)')

        if not sbilancio_assorbito.empty:
            stats = pd.merge(stats, sbilancio_assorbito, left_index=True, right_index=True, how='outer')

        stats = stats.fillna(0)
        
        # Calculation of Deltas (still in cents)
        stats['Unmatched DEBIT'] = stats['Totale Dare'] - stats['Usato Dare']
        stats['Unmatched CREDIT'] = stats['Totale Avere'] - stats['Usato Avere']
        
        # Net imbalance of only unmatched movements
        stats['Residual Imbalance (DEBIT - CREDIT)'] = stats['Unmatched DEBIT'] - stats['Unmatched CREDIT']

        # Final Monthly Imbalance
        if 'Absorbed Imbalance (in match)' not in stats.columns:
            stats['Absorbed Imbalance (in match)'] = 0
            
        stats['Final Monthly Imbalance'] = stats['Residual Imbalance (DEBIT - CREDIT)'] + stats['Absorbed Imbalance (in match)']

        # Reorganize columns for clarity
        stats = stats[[
            'Totale Dare', 'Usato Dare', 'Unmatched DEBIT',
            'Totale Avere', 'Usato Avere', 'Unmatched CREDIT',
            'Residual Imbalance (DEBIT - CREDIT)',
            'Absorbed Imbalance (in match)',
            'Final Monthly Imbalance'
        ]]

        # Sort by month
        stats = stats.sort_index()
        
        # Format the index (Period) into a string
        stats.index = stats.index.astype(str)
        stats.index.name = 'Mese'
        
        return stats.reset_index()

    def _verifica_quadratura_totali(self, tot_dare_orig, tot_avere_orig, verbose=True):
        """Verifies that the final total (used + residual) matches the original total."""
        if self.dare_df is None or self.avere_df is None:
            return

        tot_dare_final = self.dare_df['Dare'].sum()
        tot_avere_final = self.avere_df['Avere'].sum()
        
        diff_dare = tot_dare_final - tot_dare_orig
        diff_avere = tot_avere_final - tot_avere_orig
        
        if verbose:
            print("\n🔍 Verifying Total Balances (Original vs Final):")
            print(f"   DEBIT:  {tot_dare_orig/100:,.2f} € (Orig) vs {tot_dare_final/100:,.2f} € (Fin) -> Delta: {diff_dare/100:,.2f} €")
            print(f"   CREDIT: {tot_avere_orig/100:,.2f} € (Orig) vs {tot_avere_final/100:,.2f} € (Fin) -> Delta: {diff_avere/100:,.2f} €")
            
        if abs(diff_dare) > 1 or abs(diff_avere) > 1:
             print(f"⚠️  WARNING: Discrepancy detected in totals! DEBIT: {diff_dare}, CREDIT: {diff_avere}", file=sys.stderr)
        elif verbose:
             print("   ✅ Balance confirmed: No loss of amounts during splitting.")

    def _crea_foglio_manuale(self, writer):
        """Creates the 'MANUAL' sheet with the explanation of the algorithm and parameters."""
        ws = writer.book.create_sheet("MANUAL", 0) # Create as the first sheet

        # Styles
        title_font = Font(bold=True, size=14)
        header_font = Font(bold=True, size=12)
        
        # Contents
        manual_content = {}
        if self.algorithm == 'subset_sum':
            manual_content = {
                "title": "Algorithm: Subset Sum (Combination Search)",
                "description": [
                    ("General Description", "This algorithm attempts to solve the 'subset sum problem'. For each movement on one side (e.g., a deposit in CREDIT), it searches for a combination of one or more movements on the other side (e.g., receipts in DEBIT) whose sum matches the amount of the starting movement, within a given tolerance and time window."),
                    ("How it Works", "The process occurs in multiple passes:\n1. **Receipt Aggregation (Many DEBIT -> 1 CREDIT)**: Simulates the human logic of grouping multiple receipts to form a single deposit. It also looks for partial 'best fits', generating residuals.\n2. **Split Deposits (1 DEBIT -> Many CREDIT)**: Handles the less common case where a large receipt is deposited in multiple installments.\n3. **Residual Recovery**: Executes a final pass with a wider time window to try to match the remaining movements."),
                ],
                "params": [
                    ("Tolerance", f"{self.tolleranza / 100:.2f} €", "The maximum error margin accepted between the sum of the combined movements and the target amount."),
                    ("Time Window", f"{self.giorni_finestra} days", "The interval of days (before, after, or both) in which to search for candidate movements for a match."),
                    ("Max Combinations", f"{self.max_combinazioni}", "The maximum number of movements that can be combined to form a single match."),
                    ("Search Direction", f"{self.search_direction}", "Specifies whether to search for candidates only in the past ('past_only'), only in the future ('future_only'), or in both directions ('both') relative to the date of the target movement."),
                    ("Residual Analysis Threshold", f"{self.soglia_residui / 100:.2f} €", "The minimum amount a movement must have to be considered in the residual recovery pass."),
                    ("Residual Time Window", f"{self.giorni_finestra_residui} days", "The time window, usually wider, used specifically for the residual recovery pass."),
                ]
            }
        elif self.algorithm == 'progressive_balance':
            manual_content = {
                "title": "Algorithm: Progressive Balance (Continuous Balancing)",
                "description": [
                    ("General Description", "This algorithm simulates the behavior of an operator trying to balance the books by chronologically scrolling through the lists. It progressively sums receipts and deposits and, as soon as the two totals match, closes the block and starts over from zero."),
                    ("How it Works", "1. Separately sorts Receipts (Debit) and Deposits (Credit) by date.\n2. Maintains two separate progressive totals.\n3. If the Debit total is less than the Credit total, it adds the next receipt to 'catch up'.\n4. If the Credit total is lower, it adds the next deposit.\n5. When the totals are equal (zero difference), the group of accumulated movements is considered reconciled.\n6. **Reset/Force**: If the accumulated block exceeds the time window duration without balancing, it is reset (or forced if the option is active) to isolate the error and allow the reconciliation of subsequent movements."),
                    ("Ideal Use Cases", "Ideal for situations where receipts are deposited in bulk or vice versa, but without an immediate 1-to-1 correspondence. It naturally handles variable time windows between receipt and deposit."),
                ],
                "params": [
                    ("Tolerance", f"{self.tolleranza / 100:.2f} €", "The maximum error margin to consider the progressive balance 'zeroed' and thus identify a block of balanced transactions."),
                ]
            }
            
        # Addition of all common parameters to the manual report
        common_params = [
             ("Sorting Strategy", self.sorting_strategy, "Criterion used to sort movements before processing (e.g., Date)."),
             ("Search Direction", self.search_direction, "Preferred time direction for matches."),
             ("Numba Optimization", "Enabled" if self.use_numba else "Disabled", "Indicates whether the accelerated calculation engine was used."),
             ("Column Mapping", str(self.column_mapping), "Names of the columns in the original file mapped to Data/Dare/Avere."),
             ("Force Close on Timeout", "Yes" if self.ignore_tolerance else "No", "If Yes, accepts non-squared blocks if they exceed the time window.")
        ]
        if 'params' in manual_content:
            manual_content['params'].extend(common_params)
        else:
            manual_content['params'] = common_params

        # Writing to the sheet
        row_cursor = 1
        ws.cell(row=row_cursor, column=1, value=manual_content.get('title')).font = title_font
        row_cursor += 2

        for header, text in manual_content.get('description', []):
            ws.cell(row=row_cursor, column=1, value=header).font = header_font
            row_cursor += 1
            cell = ws.cell(row=row_cursor, column=1, value=text)
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            ws.merge_cells(start_row=row_cursor, start_column=1, end_row=row_cursor, end_column=5)
            row_cursor += 2
        
        ws.cell(row=row_cursor, column=1, value="Parameters Used in this Execution").font = title_font
        row_cursor += 1
        ws.cell(row=row_cursor, column=1, value="Parameter").font = header_font
        ws.cell(row=row_cursor, column=2, value="Value").font = header_font
        ws.cell(row=row_cursor, column=3, value="Meaning").font = header_font
        row_cursor += 1

        for name, value, desc in manual_content.get('params', []):
            ws.cell(row=row_cursor, column=1, value=name)
            ws.cell(row=row_cursor, column=2, value=value)
            cell = ws.cell(row=row_cursor, column=3, value=desc)
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            ws.merge_cells(start_row=row_cursor, start_column=3, end_row=row_cursor, end_column=5)
            row_cursor += 1

        # Adjust column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 80

    def _crea_report_excel(self, output_file, original_df):
        """Saves the results to a multi-sheet Excel file."""
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            # --- MANUAL SHEET (NEW) ---
            self._crea_foglio_manuale(writer)

            # --- Matches Sheet ---
            if self.df_abbinamenti is not None and not self.df_abbinamenti.empty:
                df_abbinamenti_excel = self.df_abbinamenti.copy()
                # --- CHANGE: Simplified and corrected list formatting for Excel output ---
                # Function to correctly format index lists
                def format_index_list(index_list):
                    if not isinstance(index_list, list): return index_list
                    # FIX: Adds 2 to align the 0-based pandas index with row 2 of Excel
                    return ', '.join(map(str, [i + 2 for i in index_list]))

                # Function to correctly format a single numeric value (e.g., somma_avere, differenza)
                def format_currency_value(value):
                    if pd.isna(value): return ''
                    # Divide by 100 and format with 2 decimal places, using a comma as a separator
                    return f"{value/100:.2f}".replace('.', ',')

                # Function to correctly format lists of amounts (float numbers) with a comma
                # The original function already used cents, so just replace the dot with a comma.
                # I have modified the format_list function to include the replacement of the dot with a comma.
                # The division by 100 is already present.

                def format_list(data, is_float=False):
                    if not isinstance(data, list): return data
                    items = [f"{i/100:.2f}".replace('.', ',') for i in data] if is_float else data
                    return ', '.join(map(str, items))

                for col in ['dare_indices', 'avere_indices']: df_abbinamenti_excel[col] = df_abbinamenti_excel[col].apply(format_index_list)
                for col in ['dare_importi', 'avere_importi']: df_abbinamenti_excel[col] = df_abbinamenti_excel[col].apply(lambda x: format_list(x, is_float=True))
                df_abbinamenti_excel['dare_date'] = df_abbinamenti_excel['dare_date'].apply(lambda x: ', '.join([d.strftime('%d/%m/%y') for d in x]) if isinstance(x, list) else x.strftime('%d/%m/%y'))
                df_abbinamenti_excel['avere_data'] = pd.to_datetime(df_abbinamenti_excel['avere_data']).dt.strftime('%d/%m/%y')

                # Apply formatting for somma_avere and differenza
                df_abbinamenti_excel['somma_avere'] = df_abbinamenti_excel['somma_avere'].apply(format_currency_value)
                df_abbinamenti_excel['differenza'] = df_abbinamenti_excel['differenza'].apply(format_currency_value)

                df_abbinamenti_excel.to_excel(writer, sheet_name='Matches', index=False)

                # --- ROW COLORING BY PASS (MODIFIED) ---
                ws = writer.sheets['Matches']
                
                fill_pass1 = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid") # Green
                fill_pass2 = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid") # Yellow
                fill_pass3 = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid") # Red
                fill_progressive = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid") # Light blue

                if 'pass_name' in df_abbinamenti_excel.columns:
                    for i, row in df_abbinamenti_excel.iterrows():
                        pass_name = str(row['pass_name'])
                        fill = None
                        if self.algorithm == 'subset_sum':
                            if "Pass 1" in pass_name: fill = fill_pass1
                            elif "Pass 2" in pass_name: fill = fill_pass2
                            elif "Pass 3" in pass_name: fill = fill_pass3
                        elif self.algorithm == 'progressive_balance':
                            fill = fill_progressive
                        
                        if fill:
                            excel_row = i + 2
                            for col in range(1, len(df_abbinamenti_excel.columns) + 1):
                                ws.cell(row=excel_row, column=col).fill = fill

            # --- Unreconciled Sheets ---
            if self.dare_non_util is not None and not self.dare_non_util.empty:
                df_dare_report = self.dare_non_util[['indice_orig', 'Data', 'Dare']].copy()
                # FIX: Adds 2 to align the index with the Excel row
                df_dare_report['indice_orig'] = df_dare_report['indice_orig'] + 2
                df_dare_report['Data'] = pd.to_datetime(df_dare_report['Data']).dt.strftime('%d/%m/%y')
                df_dare_report['Dare'] = df_dare_report['Dare'] / 100.0
                df_dare_report.rename(columns={'indice_orig': 'Row Index', 'Dare': 'Amount'}).to_excel(writer, sheet_name='Unused DEBIT', index=False)

            if self.avere_non_riconc is not None and not self.avere_non_riconc.empty:
                df_avere_report = self.avere_non_riconc[['indice_orig', 'Data', 'Avere']].copy()
                # FIX: Adds 2 to align the index with the Excel row
                df_avere_report['indice_orig'] = df_avere_report['indice_orig'] + 2
                df_avere_report['Data'] = pd.to_datetime(df_avere_report['Data']).dt.strftime('%d/%m/%y')
                df_avere_report['Avere'] = df_avere_report['Avere'] / 100.0
                df_avere_report.rename(columns={'indice_orig': 'Row Index', 'Avere': 'Amount'}).to_excel(writer, sheet_name='Unreconciled CREDIT', index=False)

            # --- Sheet with original data ---
            df_originale_report = original_df.copy()
            if 'Data' in df_originale_report.columns:
                 df_originale_report.sort_values(by=['Data', 'indice_orig'], inplace=True)
            if 'Dare' in df_originale_report.columns: df_originale_report['Dare'] = df_originale_report['Dare'] / 100
            if 'Avere' in df_originale_report.columns: df_originale_report['Avere'] = df_originale_report['Avere'] / 100
            if 'Dare' in df_originale_report.columns and 'Avere' in df_originale_report.columns:
                df_originale_report['Progressive Balance'] = (df_originale_report['Dare'] - df_originale_report['Avere']).cumsum()
            if 'Data' in df_originale_report.columns:
                df_originale_report['Data'] = pd.to_datetime(df_originale_report['Data']).dt.strftime('%d/%m/%Y')
                if 'indice_orig' in df_originale_report.columns and 'usato' not in df_originale_report.columns:
                    df_originale_report.drop(columns=['indice_orig'], inplace=True)
            df_originale_report.to_excel(writer, sheet_name='Originale', index=False)

            # --- Statistics Sheet ---
            stats = self.get_stats()
            if stats and self.dare_df is not None and self.avere_df is not None:
                def format_eur(value): return f"{value:,.2f} €".replace(",", "X").replace(".", ",").replace("X", ".")
                
                df_incassi = pd.DataFrame({
                    'TOT': [stats.get('Totale Incassi (DARE)'), format_eur(self.dare_df['Dare'].sum() / 100)],
                    'USATI': [stats.get('Incassi (DARE) utilizzati'), format_eur(self.dare_df[self.dare_df['usato']]['Dare'].sum() / 100)],
                    '% USATI': [stats.get('% Incassi (DARE) utilizzati (Num)'), f"{stats.get('_raw_perc_dare_importo', 0):.2f}%"],
                    'Delta': [stats.get('Incassi (DARE) non utilizzati'), format_eur(stats.get('_raw_importo_dare_non_util', 0))]
                }, index=['Numero', 'Importo'])

                df_versamenti = pd.DataFrame({
                    'TOT': [stats.get('Totale Versamenti (AVERE)'), format_eur(self.avere_df['Avere'].sum() / 100)],
                    'USATI': [stats.get('Versamenti (AVERE) riconciliati'), format_eur(self.avere_df[self.avere_df['usato']]['Avere'].sum() / 100)],
                    '% USATI': [stats.get('% Versamenti (AVERE) riconciliati (Num)'), f"{stats.get('_raw_perc_avere_importo', 0):.2f}%"],
                    'Delta': [stats.get('Versamenti (AVERE) non riconciliati'), format_eur(stats.get('_raw_importo_avere_non_riconc', 0))]
                }, index=['Numero', 'Importo'])

                delta_conteggio = stats.get('Incassi (DARE) non utilizzati', 0) - stats.get('Versamenti (AVERE) non riconciliati', 0)
                df_confronto = pd.DataFrame({
                    'Count Delta': [delta_conteggio],
                    'Amount Delta (€)': [stats.get('Final delta (DEBIT - CREDIT)')]
                }, index=['Incassi vs Versamenti'])
                
                # Add info on structural imbalance
                df_strutturale = pd.DataFrame({
                    'Info': ['Difference present in original data (DEBIT - CREDIT)'],
                    'Amount': [stats.get('Structural Imbalance (Source)')]
                })

                df_incassi.to_excel(writer, sheet_name='Statistics', startrow=2)
                df_versamenti.to_excel(writer, sheet_name='Statistics', startrow=8)
                df_confronto.to_excel(writer, sheet_name='Statistics', startrow=14)
                df_strutturale.to_excel(writer, sheet_name='Statistics', startrow=18, index=False)

            sheet_stats = writer.sheets['Statistics']
            sheet_stats.cell(row=1, column=1, value="Receipts Summary (DEBIT)")
            sheet_stats.cell(row=7, column=1, value="Deposits Summary (CREDIT)")
            sheet_stats.cell(row=13, column=1, value="Final Imbalance Comparison")
            sheet_stats.cell(row=17, column=1, value="Structural Imbalance Analysis (Initial Data)")

            # --- ADDITION: Monthly Balance Sheet ---
            df_mensile = self._calcola_quadratura_mensile()
            if not df_mensile.empty:
                # Convert from cents to Euros (float) to allow the chart to work
                cols_to_convert = [c for c in df_mensile.columns if c != 'Mese']
                for col in cols_to_convert:
                    df_mensile[col] = df_mensile[col] / 100.0
                
                # --- NEW: Helper column for the chart ---
                # To display correctly, unmatched CREDITs must be negative.
                df_mensile['Unmatched CREDIT (Chart)'] = -df_mensile['Unmatched CREDIT']
                
                df_mensile.to_excel(writer, sheet_name='Monthly Balance', index=False)
                
                ws = writer.sheets['Monthly Balance']
                
                # Apply currency formatting to cells (because they are now pure numbers)
                for col_idx, col_name in enumerate(df_mensile.columns, start=1):
                    if col_name != 'Mese' and '(Chart)' not in col_name: # Exclude helper column
                        for row in range(2, len(df_mensile) + 2):
                            cell = ws.cell(row=row, column=col_idx)
                            cell.number_format = '#,##0.00 €'
                
                # Adjust column widths
                for idx, col in enumerate(df_mensile.columns):
                    if '(Chart)' not in col: # Do not adjust the hidden column
                        max_len = min(50, max(df_mensile[col].astype(str).map(len).max() if not df_mensile.empty else 0, len(str(col)))) + 2
                        ws.column_dimensions[chr(65 + idx)].width = max_len

                # --- IMBALANCE CHART (MODIFIED) ---
                try:
                    chart = BarChart()
                    chart.type = "col"
                    chart.style = 10
                    chart.title = "Monthly Imbalance Composition"
                    chart.y_axis.title = "Amount (€)"
                    chart.x_axis.title = "Month"
                    chart.grouping = "stacked"
                    chart.overlap = 100
                    
                    # Columns: Month, Unmatched DEBIT, Unmatched CREDIT (Chart), Absorbed Imbalance
                    col_mese_idx = df_mensile.columns.get_loc('Mese') + 1
                    col_dare_idx = df_mensile.columns.get_loc('Unmatched DEBIT') + 1
                    col_avere_grafico_idx = df_mensile.columns.get_loc('Unmatched CREDIT (Chart)') + 1
                    col_assorbito_idx = df_mensile.columns.get_loc('Absorbed Imbalance (in match)') + 1
                    
                    cats = Reference(ws, min_col=col_mese_idx, min_row=2, max_row=len(df_mensile)+1)
                    
                    # Fix for openpyxl >= 3.1: Modify the cell header instead of assigning .title to the series
                    # This avoids the error: TypeError: .tx should be SeriesLabel but value is str
                    ws.cell(row=1, column=col_avere_grafico_idx, value="Unmatched CREDIT")

                    # Series 1: Unmatched DEBIT
                    data_dare = Reference(ws, min_col=col_dare_idx, min_row=1, max_row=len(df_mensile)+1)
                    chart.add_data(data_dare, titles_from_data=True)
                    
                    # Series 2: Unmatched CREDIT (negative)
                    data_avere = Reference(ws, min_col=col_avere_grafico_idx, min_row=1, max_row=len(df_mensile)+1)
                    chart.add_data(data_avere, titles_from_data=True)
                    
                    # Series 3: Absorbed Imbalance
                    data_assorbito = Reference(ws, min_col=col_assorbito_idx, min_row=1, max_row=len(df_mensile)+1)
                    chart.add_data(data_assorbito, titles_from_data=True)

                    chart.set_categories(cats)
                    chart.shape = 4
                    chart.width = 25 # Width in cm
                    chart.height = 12 # Height in cm
                    
                    # Position the chart below the data, with a couple of rows of margin
                    chart_anchor = f"A{len(df_mensile) + 4}"
                    ws.add_chart(chart, chart_anchor)
                    
                    # Hide the helper column for the chart
                    ws.column_dimensions[chr(65 + col_avere_grafico_idx - 1)].hidden = True
                except Exception as e:
                    print(f"Could not create chart: {e}")

    def get_stats(self):
        """Calculates and returns a complete dictionary of statistics."""
        if self.dare_df is None or self.avere_df is None or 'usato' not in self.dare_df.columns or 'usato' not in self.avere_df.columns: return {}

        num_dare_tot = len(self.dare_df)
        num_dare_usati = int(self.dare_df['usato'].sum()) # Now the 'usato' column exists
        imp_dare_tot = self.dare_df['Dare'].sum() # in cents
        imp_dare_usati = self.dare_df[self.dare_df['usato']]['Dare'].sum() # in cents

        num_avere_tot = len(self.avere_df)
        num_avere_usati = int(self.avere_df['usato'].sum()) # Now the 'usato' column exists
        imp_avere_tot = self.avere_df['Avere'].sum() # in cents
        imp_avere_usati = self.avere_df[self.avere_df['usato']]['Avere'].sum() # in cents

        # Recalculate dare_non_util and avere_non_riconc based on the updated 'usato' column
        importo_dare_non_util = (self.dare_non_util['Dare'].sum() / 100) if self.dare_non_util is not None and not self.dare_non_util.empty else 0
        importo_avere_non_riconc = (self.avere_non_riconc['Avere'].sum() / 100) if self.avere_non_riconc is not None and not self.avere_non_riconc.empty else 0

        sbilancio_strutturale = imp_dare_tot - imp_avere_tot

        return {
            'Total Receipts (DEBIT)': num_dare_tot,
            'Used Receipts (DEBIT)': num_dare_usati,
            '% Used Receipts (DEBIT) (Num)': f"{(num_dare_usati / num_dare_tot * 100) if num_dare_tot > 0 else 0:.1f}%",
            '% Covered Receipts (DEBIT) (Vol)': f"{(imp_dare_usati / imp_dare_tot * 100) if imp_dare_tot > 0 else 0:.1f}%",
            'Unused Receipts (DEBIT)': num_dare_tot - num_dare_usati,
            
            'Total Deposits (CREDIT)': num_avere_tot,
            'Reconciled Deposits (CREDIT)': num_avere_usati,
            '% Reconciled Deposits (CREDIT) (Num)': f"{(num_avere_usati / num_avere_tot * 100) if num_avere_tot > 0 else 0:.1f}%",
            '% Covered Deposits (CREDIT) (Vol)': f"{(imp_avere_usati / imp_avere_tot * 100) if imp_avere_tot > 0 else 0:.1f}%",
            'Unreconciled Deposits (CREDIT)': num_avere_tot - num_avere_usati,

            'Final delta (DEBIT - CREDIT)': f"{(importo_dare_non_util - importo_avere_non_riconc):,.2f} €".replace(",", "X").replace(".", ",").replace("X", "."),
            'Structural Imbalance (Source)': f"{(sbilancio_strutturale / 100):,.2f} €".replace(",", "X").replace(".", ",").replace("X", "."),
            
            # Raw values for aggregations and internal calculations
            '_raw_importo_dare_non_util': importo_dare_non_util,
            '_raw_importo_avere_non_riconc': importo_avere_non_riconc,
            '_raw_perc_dare_importo': (imp_dare_usati / imp_dare_tot * 100) if imp_dare_tot > 0 else 0,
            '_raw_perc_avere_importo': (imp_avere_usati / imp_avere_tot * 100) if imp_avere_tot > 0 else 0,
        }

    def run(self, input_file, output_file=None, verbose=True):
        """
        Executes the entire reconciliation process.

        1. Loads (or receives) the data.
        2. Separates movements into DEBIT and CREDIT.
        3. Executes the configured reconciliation algorithms.
        4. Generates statistics and reports.

        Args:
            input_file (str or pd.DataFrame): Path of the input file or an already loaded DataFrame.
            output_file (str, optional): Path where to save the Excel report. If None, does not save.
            verbose (bool): If True, prints detailed logs to the console.

        Returns:
            dict: Dictionary containing the final reconciliation statistics.
        """
        if not NUMBA_AVAILABLE and verbose:
            # Print a warning if Numba is not available
            print("\n⚠️  WARNING: 'numba' library not found. Running in non-optimized mode (slower).")
            print("   For better performance, install it with: pip install numba\n")
        try:
           # Reset used indices for each run, important for the optimizer
            self.used_dare_indices = set()
            self.used_avere_indices = set()

            # --- CHANGE: Flexible input handling ---
            # The optimizer passes a DataFrame for efficiency, main.py passes a path.
            if isinstance(input_file, pd.DataFrame):
                if verbose: print("1. Using pre-loaded DataFrame.")
                # The input is already a processed df, we use it directly
                df = input_file
            else:
                if verbose: print(f"1. Loading and validating file: {input_file}")
                df = self.carica_file(input_file)

            # Calculation of original totals for balance verification
            tot_dare_orig = df['Dare'].sum()
            tot_avere_orig = df['Avere'].sum()

            if verbose: print("2. Separating and sorting DEBIT/CREDIT movements...")
            self._separa_movimenti(df)

            if verbose: print("3. Starting reconciliation passes...")
            
            all_abbinamenti = []

            # --- ALGORITHM CHOICE ---
            # If 'all', it first runs progressive balance (for blocks) then subset sum (for residuals)
            algorithms_to_run = []
            if self.algorithm == 'all':
                algorithms_to_run = ['progressive_balance', 'subset_sum']
            elif self.algorithm == 'progressive_balance':
                algorithms_to_run = ['progressive_balance']
            else: # subset_sum o default
                algorithms_to_run = ['subset_sum']

            for algo in algorithms_to_run:
                if algo == 'progressive_balance':
                    all_abbinamenti.extend(self._riconcilia_saldo_progressivo(verbose=verbose))
                elif algo == 'subset_sum':
                    all_abbinamenti.extend(self._riconcilia_subset_sum(verbose=verbose))

            # Common finalization
            self.dare_df['usato'] = self.dare_df['indice_orig'].isin(self.used_dare_indices)
            self.avere_df['usato'] = self.avere_df['indice_orig'].isin(self.used_avere_indices)
            self.df_abbinamenti = self._finalizza_abbinamenti(all_abbinamenti)
            # Balance verification
            self._verifica_quadratura_totali(tot_dare_orig, tot_avere_orig, verbose=verbose)

            # --- CHECK STRUCTURAL IMBALANCE ---
            diff_strutturale = tot_dare_orig - tot_avere_orig
            if verbose and abs(diff_strutturale) > 100: # > 1 euro
                 print(f"\n⚖️  INITIAL DATA ANALYSIS: Structural imbalance detected!")
                 print(f"    Total DEBIT (Receipts):    {tot_dare_orig/100:,.2f} €")
                 print(f"    Total CREDIT (Deposits): {tot_avere_orig/100:,.2f} €")
                 print(f"    Difference at source:    {diff_strutturale/100:,.2f} € (This amount can never be reconciled)")

            # Calculate the dataframes of the unused, necessary for reports and statistics
            self.dare_non_util = self.dare_df[~self.dare_df['usato']].copy()
            self.avere_non_riconc = self.avere_df[~self.avere_df['usato']].copy()

            if verbose: print("4. Calculating final statistics...")
            stats = self.get_stats()

            # If an output file is provided, save the results
            if output_file:
                if verbose: print(f"5. Generating Excel report in: {output_file}")
                self._crea_report_excel(output_file, df)
                if verbose: print("✓ Excel report created successfully.")

            if verbose: print("\n🎉 Reconciliation completed successfully!")
            return stats

        except (FileNotFoundError, ValueError, IndexError) as e:
            # Handles all known errors (file not found, missing columns, corrupted file)
            print(f"\n❌ CRITICAL ERROR during processing of '{input_file}': {e}", file=sys.stderr)
            return None
        except Exception as e:
            # Handles any other unexpected error
            print(f"\n❌ UNEXPECTED ERROR: {e}", file=sys.stderr)
            import traceback
            traceback.print_exc()
            return None

# --- FUNCTION COMPILED WITH NUMBA ---
# This function lives outside the class to be correctly compiled by Numba.
# @jit is the decorator that compiles the function into machine code.
# nopython=True ensures there is no fallback to the Python interpreter, guaranteeing maximum speed.
@jit(nopython=True) # Questo decoratore sarà quello reale di Numba o quello fittizio
def _numba_find_combination(target, candidati_np, max_combinazioni, tolleranza):
    """
    Finds a combination of candidates whose sum approaches the target.
    This version is optimized for Numba and operates on NumPy arrays.

    Args:
        target (int): The amount to reach.
        candidati_np (np.array): 2D array where each row is [amount, original_index].
        max_combinazioni (int): Maximum number of elements in the combination.
        tolleranza (int): Acceptable error margin for the sum.

    Returns:
        np.array: An array of the original indices of the found combination, or an empty array.
    """
    # Stack: (candidate_index, current_sum, level)
    # Initialize with first-level candidates
    stack = []
    n_candidati = len(candidati_np)
    
    # We iterate in reverse order to push onto the stack, so we process the largest candidates first (index 0)
    for i in range(n_candidati - 1, -1, -1):
        val = candidati_np[i, 0]
        if val <= target + tolleranza:
            stack.append((i, val, 1))
            
    path = np.full(max_combinazioni, -1, dtype=np.int64)

    while len(stack) > 0:
        idx, current_sum, level = stack.pop()
        path[level-1] = idx
        
        # Check exact match
        if abs(target - current_sum) <= tolleranza:
             result_indices = np.full(level, 0, dtype=np.int64)
             for k in range(level):
                 result_indices[k] = candidati_np[path[k], 1]
             return result_indices
             
        if level >= max_combinazioni:
            continue
            
        # Pruning: If even by adding the largest remaining values we don't reach the target
        remaining_slots = max_combinazioni - level
        if idx + 1 < n_candidati:
            # Optimistic estimate: we use the next largest value for all remaining slots
            max_add = candidati_np[idx+1, 0] * remaining_slots
            if current_sum + max_add < target - tolleranza:
                continue
        elif current_sum < target - tolleranza:
            # No candidates left and we are not at the target
            continue

        # Generate children: try subsequent candidates
        # Push in reverse order (from smallest to largest) to explore the large ones first
        for i in range(n_candidati - 1, idx, -1):
            val = candidati_np[i, 0]
            new_sum = current_sum + val
            if new_sum <= target + tolleranza:
                 stack.append((i, new_sum, level + 1))

    # If the stack becomes empty, no combination was found.
    return np.empty(0, dtype=np.int64)

@jit(nopython=True)
def _numba_find_best_fit_combination(target, candidati_np, max_combinazioni, tolleranza):
    """
    Finds the combination of candidates that maximizes the sum <= target (Best Fit / Knapsack).
    It does not look for the exact sum, but the one that comes closest without exceeding the target.
    """
    # Stack: (candidate_index, current_sum, level)
    stack = []
    n_candidati = len(candidati_np)
    
    for i in range(n_candidati - 1, -1, -1):
        val = candidati_np[i, 0]
        if val <= target + tolleranza:
            stack.append((i, val, 1))
    
    path = np.full(max_combinazioni, -1, dtype=np.int64)
    
    # Variables to track the best solution found so far
    best_sum = 0
    best_path_len = 0
    best_path = np.full(max_combinazioni, -1, dtype=np.int64)
    
    # Minimum threshold to consider a best fit useful (e.g., fill at least 1% of the target)
    min_fill_threshold = target * 0.01

    while len(stack) > 0:
        idx, current_sum, level = stack.pop()
        path[level-1] = idx
        
        # If the current sum is better than the one found so far, update the best fit
        if current_sum > best_sum:
            best_sum = current_sum
            best_path_len = level
            # Copy the current path to best_path
            for k in range(level):
                best_path[k] = path[k]
                
            # If we found an almost perfect match (within tolerance), we can stop
            if abs(target - best_sum) <= tolleranza:
                break

        if level >= max_combinazioni:
            continue
            
        remaining_slots = max_combinazioni - level
        
        # Pruning Upper Bound
        if idx + 1 < n_candidati:
             max_potential = current_sum + candidati_np[idx+1, 0] * remaining_slots
             if max_potential <= best_sum:
                 continue
        else:
             continue

        for i in range(n_candidati - 1, idx, -1):
            val = candidati_np[i, 0]
            new_sum = current_sum + val
            
            if new_sum > target + tolleranza:
                continue
                
            # Local Pruning
            if new_sum + (val * (remaining_slots - 1)) <= best_sum:
                continue
                
            stack.append((i, new_sum, level + 1))

    # If we found a valid solution
    if best_path_len > 0 and best_sum >= min_fill_threshold:
        result_indices = np.full(best_path_len, 0, dtype=np.int64)
        for k in range(best_path_len):
            result_indices[k] = candidati_np[best_path[k], 1]
        return result_indices

    return np.empty(0, dtype=np.int64)