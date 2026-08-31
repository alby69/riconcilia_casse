import calendar

import pandas as pd
from openpyxl.styles import PatternFill, Alignment, Font, NamedStyle
from openpyxl.formatting.rule import DataBarRule, ColorScaleRule, CellIsRule
from openpyxl.chart import BarChart, Reference, Series


MONTH_NAMES_IT = (
    "Gennaio", "Febbraio", "Marzo", "Aprile", "Maggio", "Giugno",
    "Luglio", "Agosto", "Settembre", "Ottobre", "Novembre", "Dicembre",
)


GROUP_FILLS = (
    ("DDEBF7", "9DC3E6"),  # Blue   (light fill / medium fill for Difference)
    ("C6EFCE", "A9D08E"),  # Green
    ("FFD966", "E6B800"),  # Amber
)
GROUP_NAMES = ("Blu", "Verde", "Arancione")


class ExcelReporter:
    """
    Manages the generation of Excel reports for the reconciliation engine.
    Separates presentation logic from business logic.
    """

    def __init__(self, engine):
        self.engine = engine
        self.currency_style = self._create_styles()
        self.original_df = None

    @staticmethod
    def _fmt_eur(value):
        """Formats a value in Euros using the Italian decimal separator."""
        if value is None:
            return ""
        s = f"{value:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
        return f"{s} €"

    def _analyze_saldo_prog(self, df):
        """Analyzes the 'Saldo Prog.' column (running cash balance of the store).

        Saldo Prog. is a running balance computed on the source rows:
            Saldo Prog. = opening_cash + cumsum(Dare - Avere)

        The analysis validates that the declared progressive balance matches the
        theoretical cash position and exposes the opening/closing cash amounts,
        useful as a cross-check (quadratura) for the reconciliation.

        Returns:
            dict: {'present': False} if the column is absent/unusable, otherwise
                  opening/closing/min cash, negative/inconsistent row counts and
                  a check_map {orig_index: message} aligned to the source rows.
        """
        result = {"present": False}
        if df is None or df.empty or "Saldo Prog." not in df.columns:
            return result

        saldo = pd.to_numeric(df["Saldo Prog."], errors="coerce")
        if saldo.notna().sum() == 0 or pd.isna(saldo.iloc[0]):
            return result

        debit_cents = df["Debit"].astype("float")
        credit_cents = df["Credit"].astype("float")
        saldo_cents = (saldo * 100).round().astype("int64")

        opening_cents = int(
            saldo_cents.iloc[0] - (debit_cents.iloc[0] - credit_cents.iloc[0])
        )
        expected_cents = opening_cents + (debit_cents - credit_cents).cumsum()
        mismatch_cents = (expected_cents - saldo_cents).abs()
        inconsistent = mismatch_cents > 1

        check_map = {}
        for pos, oi in enumerate(df["orig_index"]):
            oi_key = int(oi)
            if inconsistent.iloc[pos]:
                check_map[oi_key] = (
                    f"⚠️ Saldo incoerente (Δ {self._fmt_eur(mismatch_cents.iloc[pos] / 100.0)})"
                )
            else:
                check_map[oi_key] = ""

        result.update(
            {
                "present": True,
                "opening": opening_cents / 100.0,
                "closing": float(saldo_cents.iloc[-1]) / 100.0,
                "min": float(saldo.min()),
                "negative_rows": int((saldo < 0).sum()),
                "inconsistent_rows": int(inconsistent.sum()),
                "check_map": check_map,
            }
        )
        return result

    def _build_match_groups(self):
        """Builds a lookup {orig_index: group info} from the engine's matches.

        Each match (row of matches_df) defines a group; the fill color cycles
        every 3 groups. Both the matched DEBIT and CREDIT rows of a group share
        the same color, reproducing the D(..)_A(..) anchoring shown in Matches.

        Because the Progressive Balance algorithm can split one receipt across
        several deposits, a single orig_index may belong to multiple groups.
        The 'memberships' list records, in Matches order, the amount of that
        row actually consumed by each group (used/residual amounts). The row
        takes the color of the FIRST group that contains it (in Matches order).
        """
        membership = {}
        if self.engine.matches_df is None or self.engine.matches_df.empty:
            return membership

        for group_idx, (_, row) in enumerate(
            self.engine.matches_df.iterrows(), start=1
        ):
            color_idx = (group_idx - 1) % len(GROUP_FILLS)
            transaction_id = row.get("Transaction ID", "")
            difference = row.get("difference", 0) or 0
            for oi, amt in zip(
                row.get("debit_indices", []) or [],
                row.get("debit_amounts", []) or [],
            ):
                membership.setdefault(int(oi), []).append(
                    {
                        "group_id": group_idx,
                        "color_idx": color_idx,
                        "transaction_id": transaction_id,
                        "difference": difference,
                        "side": "debit",
                        "amount": int(amt),
                    }
                )
            for oi, amt in zip(
                row.get("credit_indices", []) or [],
                row.get("credit_amounts", []) or [],
            ):
                membership.setdefault(int(oi), []).append(
                    {
                        "group_id": group_idx,
                        "color_idx": color_idx,
                        "transaction_id": transaction_id,
                        "difference": difference,
                        "side": "credit",
                        "amount": int(amt),
                    }
                )

        groups = {}
        for oi, infos in membership.items():
            primary = infos[0]  # First group wins (Matches order)
            groups[oi] = {
                "group_id": primary["group_id"],
                "color_idx": primary["color_idx"],
                "transaction_id": primary["transaction_id"],
                "difference": primary["difference"],
                "side": primary["side"],
                "split": len(infos) > 1,
                "memberships": infos,
                "other_transaction_ids": [i["transaction_id"] for i in infos[1:]],
            }
        return groups

    def _add_original_legend(self, ws, n_rows, saldo_analysis):
        """Writes an explanatory legend below the data of the 'Original' sheet."""
        row = n_rows + 3
        ws.cell(
            row=row,
            column=1,
            value="Legenda: le celle Debit/Credit con lo stesso colore appartengono allo stesso gruppo di "
            "abbinamento (vedi foglio 'Matches'). Il colore si ripete ogni 3 gruppi.",
        ).font = Font(bold=True, color="1F4E78")
        row += 1
        for idx in range(len(GROUP_FILLS)):
            cell = ws.cell(row=row, column=1 + idx)
            cell.value = f"Gruppo {idx + 1} (mod 3) - {GROUP_NAMES[idx]}"
            cell.fill = PatternFill(
                start_color=GROUP_FILLS[idx][0],
                end_color=GROUP_FILLS[idx][0],
                fill_type="solid",
            )
        row += 2
        ws.cell(
            row=row,
            column=1,
            value="Colonna 'Difference': Δ in € del gruppo (0 = gruppo quadrato; > 0 = differenza/residuo da verificare).",
        )
        row += 1
        ws.cell(
            row=row,
            column=1,
            value="Righe inserite: se un incasso (Debit) è ripartito su più versamenti, la riga originale mostra la quota "
            "consumata dal PRIMO versamento (in ordine del foglio 'Matches'); sotto di essa viene inserita una nuova riga "
            "(stessa data) per ogni quota residua, ciascuna con il colore del proprio gruppo.",
        )

    def _create_styles(self):
        """Creates named styles for reuse in the workbook."""
        currency_style = NamedStyle(name="currency_style", number_format="#,##0.00 €")
        return currency_style

    def generate_report(self, output_file, original_df):
        self.original_df = original_df
        # Remap Transaction IDs / displayed indices to the split Original rows
        self._remap_matches_to_sheet_rows(original_df)
        with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
            # Register styles
            writer.book.add_named_style(self.currency_style)

            self._create_summary_sheet(writer)
            self._create_manual_sheet(writer)
            self._create_matches_sheet(writer)
            self._create_anomalies_sheet(writer)
            self._create_unreconciled_sheets(writer)
            self._create_original_sheet(writer, original_df)
            self._create_monthly_balance_sheet(writer)

            # Set Summary as the active sheet
            writer.book.active = writer.book["Summary"]

    def _create_summary_sheet(self, writer):
        """Creates the main 'Summary' sheet with KPIs and actionable insights."""
        ws = writer.book.create_sheet("Summary", 0)
        title_font = Font(bold=True, size=16, color="1F4E78")
        header_font = Font(bold=True, size=12, color="1F4E78")

        ws.cell(row=1, column=1, value="Reconciliation Summary").font = title_font

        # --- KPIs ---
        stats = self.engine.get_stats()
        debit_coverage = stats.get("_raw_debit_amount_perc", 0)
        credit_coverage = stats.get("_raw_credit_amount_perc", 0)

        ws.cell(row=3, column=1, value="Key Performance Indicators").font = header_font
        kpis = [
            ("Debit Coverage (Volume)", f"{debit_coverage:.2f}%"),
            ("Credit Coverage (Volume)", f"{credit_coverage:.2f}%"),
            ("Unreconciled Debits", stats.get("Unused Receipts (DEBIT)")),
            ("Unreconciled Credits", stats.get("Unreconciled Deposits (CREDIT)")),
            ("Final Delta", stats.get("Final delta (DEBIT - CREDIT)")),
        ]

        # --- Saldo Prog. (cash balance) KPIs: useful for the quadratura ---
        saldo = self._analyze_saldo_prog(self.original_df)
        if saldo.get("present"):
            kpis.append(("Saldo Prog. Iniziale (Cassa)", self._fmt_eur(saldo["opening"])))
            kpis.append(("Saldo Prog. Finale (Cassa)", self._fmt_eur(saldo["closing"])))
            kpis.append(
                (
                    "Variazione Cassa (Dare - Avere)",
                    self._fmt_eur(saldo["closing"] - saldo["opening"]),
                )
            )
            if saldo["negative_rows"] > 0:
                kpis.append(("⚠ Righe con Saldo Negativo", saldo["negative_rows"]))
            if saldo["inconsistent_rows"] > 0:
                kpis.append(("⚠ Righe Saldo Incoerenti", saldo["inconsistent_rows"]))

        for i, (label, value) in enumerate(kpis, 4):
            ws.cell(row=i, column=1, value=label).font = Font(bold=True)
            ws.cell(row=i, column=2, value=value)

        cursor = 4 + len(kpis) + 1

        # --- Textual Summary ---
        ws.cell(row=cursor, column=1, value="Automated Analysis").font = header_font
        summary_text = f"Reconciliation resulted in {debit_coverage:.1f}% of debit volume and {credit_coverage:.1f}% of credit volume being matched. "
        if saldo.get("present"):
            summary_text += (
                f"Cash balance (Saldo Prog.): {self._fmt_eur(saldo['opening'])} at start, "
                f"{self._fmt_eur(saldo['closing'])} at end. "
            )
            if saldo["inconsistent_rows"] > 0:
                summary_text += (
                    f"WARNING: {saldo['inconsistent_rows']} rows show an inconsistent progressive balance. "
                )
            elif saldo["negative_rows"] > 0:
                summary_text += (
                    f"WARNING: {saldo['negative_rows']} rows show a negative cash balance. "
                )
        if debit_coverage > 95 and credit_coverage > 95:
            summary_text += "This is a great result, with very few items left to check."
        elif debit_coverage < 80 or credit_coverage < 80:
            summary_text += "There is a significant amount of unreconciled transactions. Focus on the largest unmatched items listed below."
        else:
            summary_text += "Good result, but some items require manual review."

        text_row = cursor + 1
        ws.cell(row=text_row, column=1, value=summary_text).alignment = Alignment(
            wrap_text=True
        )
        ws.merge_cells(f"A{text_row}:E{text_row}")

        # --- Top 5 Unreconciled Items ---
        top_row = cursor + 4
        ws.cell(
            row=top_row, column=1, value="Top 5 Largest Unreconciled Debits"
        ).font = header_font
        if (
            self.engine.unused_debit_df is not None
            and not self.engine.unused_debit_df.empty
        ):
            top_debits = self.engine.unused_debit_df.nlargest(5, "Debit")
            ws.cell(row=top_row + 1, column=1, value="Date").font = Font(bold=True)
            ws.cell(row=top_row + 1, column=2, value="Amount").font = Font(bold=True)
            for i, row in enumerate(top_debits.itertuples(), top_row + 2):
                ws.cell(row=i, column=1, value=row.Date.strftime("%d/%m/%Y"))
                cell = ws.cell(row=i, column=2, value=row.Debit / 100)
                cell.style = "currency_style"

        ws.cell(
            row=top_row, column=4, value="Top 5 Largest Unreconciled Credits"
        ).font = header_font
        if (
            self.engine.unreconciled_credit_df is not None
            and not self.engine.unreconciled_credit_df.empty
        ):
            top_credits = self.engine.unreconciled_credit_df.nlargest(5, "Credit")
            ws.cell(row=top_row + 1, column=4, value="Date").font = Font(bold=True)
            ws.cell(row=top_row + 1, column=5, value="Amount").font = Font(bold=True)
            for i, row in enumerate(top_credits.itertuples(), top_row + 2):
                ws.cell(row=i, column=4, value=row.Date.strftime("%d/%m/%Y"))
                cell = ws.cell(row=i, column=5, value=row.Credit / 100)
                cell.style = "currency_style"

        # Column widths
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 15
        ws.column_dimensions["C"].width = 5
        ws.column_dimensions["D"].width = 25
        ws.column_dimensions["E"].width = 15

    def _create_manual_sheet(self, writer):
        ws = writer.book.create_sheet("MANUAL", 1)
        # ... (rest of the logic is similar, just update the content)
        title_font = Font(bold=True, size=14)
        header_font = Font(bold=True, size=12)

        algo_map = {
            "subset_sum": {
                "title": "Algorithm: Subset Sum (Combination Search)",
                "description": [
                    (
                        "How it Works",
                        "1. Receipt Aggregation (Many DEBIT -> 1 CREDIT)\n2. Split Deposits (1 DEBIT -> Many CREDIT)\n3. Residual Recovery",
                    )
                ],
            },
            "progressive_balance": {
                "title": "Algorithm: Progressive Balance (Sequential)",
                "description": [
                    (
                        "How it Works",
                        "Simulates an operator scrolling through lists and closing blocks when totals match.",
                    )
                ],
            },
            "greedy_amount_first": {
                "title": "Algorithm: Greedy Amount First",
                "description": [
                    (
                        "How it Works",
                        "Sorts all transactions by amount and tries to match the largest items first, which is useful for finding key matches quickly.",
                    )
                ],
            },
        }
        manual_content = algo_map.get(
            self.engine.algorithm,
            {"title": f"Algorithm: {self.engine.algorithm}", "description": []},
        )

        params = [
            (
                "Tolerance",
                f"{self.engine.tolerance / 100:.2f} €",
                "Maximum error margin.",
            ),
            (
                "Time Window",
                f"{self.engine.days_window} days",
                "Search interval for matches.",
            ),
            (
                "Max Combinations",
                f"{self.engine.max_combinations}",
                "Max elements in a combination (for Subset Sum).",
            ),
            (
                "Search Direction",
                self.engine.search_direction,
                "Time direction for search.",
            ),
            (
                "Force Close on Timeout",
                "Yes" if self.engine.ignore_tolerance else "No",
                "For Progressive Balance: accepts non-squared blocks on timeout.",
            ),
        ]
        manual_content["params"] = params

        row_cursor = 1
        ws.cell(
            row=row_cursor, column=1, value=manual_content.get("title")
        ).font = title_font
        row_cursor += 2

        for header, text in manual_content.get("description", []):
            ws.cell(row=row_cursor, column=1, value=header).font = header_font
            row_cursor += 1
            cell = ws.cell(row=row_cursor, column=1, value=text)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            ws.merge_cells(
                start_row=row_cursor, start_column=1, end_row=row_cursor, end_column=5
            )
            row_cursor += 2

        ws.cell(row=row_cursor, column=1, value="Parameters Used").font = title_font
        row_cursor += 1
        for name, value, desc in manual_content.get("params", []):
            ws.cell(row=row_cursor, column=1, value=name).font = header_font
            ws.cell(row=row_cursor, column=2, value=value)
            ws.cell(row=row_cursor, column=3, value=desc)
            row_cursor += 1

        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 20
        ws.column_dimensions["C"].width = 60

    def _create_matches_sheet(self, writer):
        if self.engine.matches_df is None or self.engine.matches_df.empty:
            return
        df = self.engine.matches_df.copy()

        # Use the split-sheet row numbers for the displayed indices when the
        # reporter remapped them (generate_report always does).
        if "debit_display" in df.columns:
            df["debit_indices"] = df["debit_display"]
        if "credit_display" in df.columns:
            df["credit_indices"] = df["credit_display"]
        df.drop(columns=["debit_display", "credit_display"], errors="ignore", inplace=True)

        def format_list(data, is_float=False):
            if not isinstance(data, list):
                return data
            items = [
                f"{i / 100:.2f}".replace(".", ",") if is_float else str(i)
                for i in data
            ]
            return ", ".join(items)

        df["debit_indices"] = df["debit_indices"].apply(format_list)
        df["credit_indices"] = df["credit_indices"].apply(format_list)
        df["debit_amounts"] = df["debit_amounts"].apply(
            lambda x: format_list(x, is_float=True)
        )
        df["credit_amounts"] = df["credit_amounts"].apply(
            lambda x: format_list(x, is_float=True)
        )

        if "total_debit" in df.columns:
            df["total_debit"] = df["total_debit"].fillna(0) / 100

        df["debit_dates"] = df["debit_dates"].apply(
            lambda x: ", ".join([d.strftime("%d/%m/%y") for d in x])
            if isinstance(x, list) and x
            else (x.strftime("%d/%m/%y") if pd.notna(x) else "")
        )
        df["credit_date"] = pd.to_datetime(df["credit_date"]).dt.strftime("%d/%m/%y")
        df["total_credit"] = df["total_credit"] / 100
        df["difference"] = df["difference"].fillna(0).astype(float) / 100

        df.to_excel(writer, sheet_name="Matches", index=False)
        ws = writer.sheets["Matches"]

        for c_idx, col_name in enumerate(df.columns, 1):
            if col_name in [
                "total_credit",
                "total_debit",
                "difference",
                "debit_amounts",
                "credit_amounts",
            ]:
                for r_idx in range(2, len(df) + 2):
                    ws.cell(row=r_idx, column=c_idx).style = self.currency_style
            elif col_name in ["days_diff", "num_credits"]:
                for r_idx in range(2, len(df) + 2):
                    ws.cell(row=r_idx, column=c_idx).number_format = "0"

        fills = {
            "Pass 1": PatternFill(
                start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"
            ),
            "Pass 2": PatternFill(
                start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"
            ),
            "Pass 3": PatternFill(
                start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"
            ),
            "Progressive": PatternFill(
                start_color="DDEBF7", end_color="DDEBF7", fill_type="solid"
            ),
            "Greedy": PatternFill(
                start_color="E9D8F5", end_color="E9D8F5", fill_type="solid"
            ),
            "Forced": PatternFill(
                start_color="FF6B6B", end_color="FF6B6B", fill_type="solid"
            ),
            "Residual Recovery": PatternFill(
                start_color="4ECDC4", end_color="4ECDC4", fill_type="solid"
            ),
        }
        if "pass_name" in df.columns:
            for i, row in df.iterrows():
                pass_name = str(row["pass_name"])
                for key, fill in fills.items():
                    if key in pass_name:
                        for col in range(1, len(df.columns) + 1):
                            ws.cell(row=i + 2, column=col).fill = fill
                        break

    def _create_anomalies_sheet(self, writer):
        """Exports only the ANOMALY matches to a dedicated 'Anomalie' sheet.

        Columns mirror the 'Matches' sheet (with the split-sheet row numbers)
        plus the uncovered amount, so every disallineamento can be verified.
        """
        if self.engine.matches_df is None or self.engine.matches_df.empty:
            return
        m = self.engine.matches_df.copy()
        if "match_type" not in m.columns:
            return
        anom = m[m["match_type"].astype(str).str.startswith("ANOMALY")].copy()
        if anom.empty:
            return

        if "debit_display" in anom.columns:
            anom["debit_indices"] = anom["debit_display"]
        if "credit_display" in anom.columns:
            anom["credit_indices"] = anom["credit_display"]
        anom.drop(
            columns=["debit_display", "credit_display"], errors="ignore", inplace=True
        )

        def format_list(data, is_float=False):
            if not isinstance(data, list):
                return data
            items = [
                f"{i / 100:.2f}".replace(".", ",") if is_float else str(i)
                for i in data
            ]
            return ", ".join(items)

        anom["debit_indices"] = anom["debit_indices"].apply(format_list)
        anom["credit_indices"] = anom["credit_indices"].apply(format_list)
        anom["debit_amounts"] = anom["debit_amounts"].apply(
            lambda x: format_list(x, is_float=True)
        )
        anom["credit_amounts"] = anom["credit_amounts"].apply(
            lambda x: format_list(x, is_float=True)
        )
        anom["total_debit"] = anom["total_debit"].fillna(0) / 100
        anom["debit_dates"] = anom["debit_dates"].apply(
            lambda x: ", ".join([d.strftime("%d/%m/%y") for d in x])
            if isinstance(x, list) and x
            else (x.strftime("%d/%m/%y") if pd.notna(x) else "")
        )
        anom["credit_date"] = pd.to_datetime(anom["credit_date"]).dt.strftime("%d/%m/%y")
        anom["total_credit"] = anom["total_credit"] / 100
        anom["difference"] = anom["difference"].fillna(0).astype(float) / 100
        anom["uncovered"] = anom["difference"]

        keep = [
            "Transaction ID",
            "credit_date",
            "total_credit",
            "debit_indices",
            "debit_dates",
            "debit_amounts",
            "total_debit",
            "uncovered",
            "match_type",
        ]
        anom = anom[[c for c in keep if c in anom.columns]]

        anom.to_excel(writer, sheet_name="Anomalie", index=False)
        ws = writer.sheets["Anomalie"]

        for c_idx, col_name in enumerate(anom.columns, 1):
            if col_name in ["total_credit", "total_debit", "uncovered", "debit_amounts"]:
                for r_idx in range(2, len(anom) + 2):
                    ws.cell(row=r_idx, column=c_idx).style = self.currency_style

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(
            start_color="C00000", end_color="C00000", fill_type="solid"
        )
        for c_idx in range(1, len(anom.columns) + 1):
            cell = ws.cell(row=1, column=c_idx)
            cell.font = header_font
            cell.fill = header_fill

        for row in ws.iter_rows(min_row=2, min_col=1, max_col=len(anom.columns)):
            for cell in row:
                cell.fill = PatternFill(
                    start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"
                )

        widths = {
            "Transaction ID": 22,
            "credit_date": 12,
            "total_credit": 14,
            "debit_indices": 24,
            "debit_dates": 28,
            "debit_amounts": 30,
            "total_debit": 14,
            "uncovered": 14,
            "match_type": 48,
        }
        for c_idx, col_name in enumerate(anom.columns, 1):
            ws.column_dimensions[
                ws.cell(row=1, column=c_idx).column_letter
            ].width = widths.get(col_name, 16)

    def _create_unreconciled_sheets(self, writer):
        # ... identical logic to before, just applying currency style
        if (
            self.engine.unused_debit_df is not None
            and not self.engine.unused_debit_df.empty
        ):
            df = self.engine.unused_debit_df[["orig_index", "Date", "Debit"]].copy()
            df.rename(
                columns={"orig_index": "Row Index", "Debit": "Amount"}, inplace=True
            )
            df["Row Index"] += 2
            df["Date"] = pd.to_datetime(df["Date"]).dt.strftime("%d/%m/%y")
            df["Amount"] /= 100.0
            df.to_excel(writer, sheet_name="Unused DEBIT", index=False)
            ws = writer.sheets["Unused DEBIT"]
            for row in ws.iter_rows(
                min_row=2, max_row=len(df) + 1, min_col=3, max_col=3
            ):
                for cell in row:
                    cell.style = self.currency_style

        if (
            self.engine.unreconciled_credit_df is not None
            and not self.engine.unreconciled_credit_df.empty
        ):
            df = self.engine.unreconciled_credit_df[
                ["orig_index", "Date", "Credit"]
            ].copy()
            df.rename(
                columns={"orig_index": "Row Index", "Credit": "Amount"}, inplace=True
            )
            df["Row Index"] += 2
            df["Date"] = pd.to_datetime(df["Date"]).dt.strftime("%d/%m/%y")
            df["Amount"] /= 100.0
            df.to_excel(writer, sheet_name="Unreconciled CREDIT", index=False)
            ws = writer.sheets["Unreconciled CREDIT"]
            for row in ws.iter_rows(
                min_row=2, max_row=len(df) + 1, min_col=3, max_col=3
            ):
                for cell in row:
                    cell.style = self.currency_style

    def _build_original_layout(self, original_df):
        """Builds the expanded layout of the 'Original' sheet.

        Returns a dict with:
          - expanded:  list of row Series (original rows + inserted residual rows
                       + a monthly subtotal row at each month change)
          - meta:      parallel per-row metadata (color/side/difference/group...)
          - portion_rows: {orig_index: [excel data rows]} in memberships order
          - groups, saldo_analysis, n_rows
        """
        df = original_df.copy()
        saldo_analysis = self._analyze_saldo_prog(df)
        check_map = saldo_analysis.get("check_map", {})

        if "Date" in df.columns:
            df.sort_values(by=["Date", "orig_index"], inplace=True)

        groups = self._build_match_groups()

        expanded = []
        meta = []
        portion_rows = {}
        excel_row = 1  # first data row is Excel row 2 (header is row 1)

        # --- monthly subtotal tracking (inserted at each month change) ---
        cols = list(df.columns)
        month_key = None
        month_deb = month_cre = 0
        month_dates = []

        def flush_month_total():
            nonlocal excel_row, month_deb, month_cre, month_dates
            if month_dates:
                end = max(month_dates)
                total_date = end.replace(
                    day=calendar.monthrange(end.year, end.month)[1]
                )
                label = (
                    "TOTALE MESE "
                    f"{MONTH_NAMES_IT[end.month - 1]} {end.year}"
                )
                total_row = pd.Series([None] * len(cols), index=cols)
                total_row["Date"] = total_date
                total_row["Debit"] = month_deb
                total_row["Credit"] = month_cre
                expanded.append(total_row)
                excel_row += 1
                meta.append(
                    {
                        "month_total": True,
                        "group_label": label,
                        "color_idx": None,
                        "side": None,
                        "difference": month_deb - month_cre,
                        "check": "",
                        "inserted": False,
                    }
                )
            month_deb = month_cre = 0
            month_dates = []

        for _, r in df.iterrows():
            oi = r["orig_index"]
            key = int(oi) if pd.notna(oi) else None

            date = r.get("Date")
            if pd.notna(date):
                d = pd.Timestamp(date)
                cur_month = (d.year, d.month)
                if month_key is not None and cur_month != month_key:
                    flush_month_total()
                month_key = cur_month
                month_dates.append(d)
                month_deb += int(r.get("Debit") or 0)
                month_cre += int(r.get("Credit") or 0)

            info = groups.get(key)

            if info and info["side"] == "debit":
                members = info["memberships"]
                portions = [(m["amount"], m) for m in members]
                leftover = int(r["Debit"]) - sum(a for a, _ in portions)
                if leftover > 0:
                    portions.append((leftover, None))

                if len(portions) == 1:
                    amount, member = portions[0]
                    row = r.copy()
                    row["Debit"] = amount
                    expanded.append(row)
                    excel_row += 1
                    meta.append(
                        self._row_meta(member, check_map.get(key, ""), False)
                    )
                    portion_rows[key] = [excel_row]
                else:
                    rows_this = []
                    for i, (amount, member) in enumerate(portions):
                        row = r.copy()
                        row["Debit"] = amount
                        inserted = i > 0
                        if inserted and "Saldo Prog." in row.index:
                            row["Saldo Prog."] = None
                        expanded.append(row)
                        excel_row += 1
                        meta.append(
                            self._row_meta(
                                member,
                                "" if inserted else check_map.get(key, ""),
                                inserted,
                            )
                        )
                        rows_this.append(excel_row)
                    portion_rows[key] = rows_this
            else:
                expanded.append(r.copy())
                excel_row += 1
                meta.append(self._row_meta(info, check_map.get(key, ""), False))
                if key is not None:
                    portion_rows[key] = [excel_row]

        flush_month_total()

        return {
            "expanded": expanded,
            "meta": meta,
            "portion_rows": portion_rows,
            "groups": groups,
            "saldo_analysis": saldo_analysis,
            "n_rows": len(expanded),
        }

    def _remap_matches_to_sheet_rows(self, original_df):
        """Rewrites Transaction IDs / displayed indices to the row numbers of the
        SPLIT 'Original' sheet.

        The engine builds D(..)_A(..) IDs from the ORIGINAL row indices. After a
        receipt is split into extra rows, those indices no longer match the rows
        the operator sees. This method maps each orig_index (and each split
        portion) to its final row number in the 'Original' sheet and updates
        matches_df accordingly (Transaction ID + debit_display/credit_display).
        """
        if self.engine.matches_df is None or self.engine.matches_df.empty:
            return

        layout = self._build_original_layout(original_df)
        groups = layout["groups"]
        portion_rows = layout["portion_rows"]

        # orig_index -> {group_id: position in memberships}
        gid_to_pos = {}
        for oi, info in groups.items():
            for pos, m in enumerate(info["memberships"]):
                gid_to_pos.setdefault(oi, {})[m["group_id"]] = pos

        matches = self.engine.matches_df
        debit_display = []
        credit_display = []
        tids = []
        for gid, (_, row) in enumerate(matches.iterrows(), start=1):
            d_rows = []
            for d in row.get("debit_indices", []) or []:
                d = int(d)
                pos = gid_to_pos.get(d, {}).get(gid)
                rows = portion_rows.get(d, [])
                if rows:
                    d_rows.append(rows[pos if pos is not None and pos < len(rows) else 0])
            c_rows = []
            for c in row.get("credit_indices", []) or []:
                rows = portion_rows.get(int(c), [])
                if rows:
                    c_rows.append(rows[0])
            debit_display.append(d_rows)
            credit_display.append(c_rows)
            tids.append(
                "D({})_A({})".format(
                    ",".join(map(str, d_rows)), ",".join(map(str, c_rows))
                )
            )

        matches["debit_display"] = debit_display
        matches["credit_display"] = credit_display
        matches["Transaction ID"] = tids

    def _create_original_sheet(self, writer, original_df):
        layout = self._build_original_layout(original_df)
        expanded = layout["expanded"]
        meta = layout["meta"]
        saldo_analysis = layout["saldo_analysis"]

        df_out = pd.DataFrame(expanded).reset_index(drop=True)

        # Attach group / delta to each (expanded) row
        df_out["Gruppo"] = [m["group_label"] for m in meta]
        df_out["Difference"] = [
            (m["difference"] or 0) / 100.0 if m["difference"] is not None else None
            for m in meta
        ]

        # Keep only the essential columns in the Original sheet
        keep_cols = ["Date", "Debit", "Credit"]
        if "Saldo Prog." in df_out.columns:
            keep_cols.append("Saldo Prog.")
        keep_cols += ["Gruppo", "Difference"]
        df_out = df_out[[c for c in keep_cols if c in df_out.columns]]

        if "Debit" in df_out.columns:
            df_out["Debit"] = df_out["Debit"] / 100
        if "Credit" in df_out.columns:
            df_out["Credit"] = df_out["Credit"] / 100
        # Format the Date column as GG/MM/AAAA
        for col in df_out.select_dtypes(include=["datetime64"]).columns:
            df_out[col] = pd.to_datetime(df_out[col], errors="coerce").dt.strftime(
                "%d/%m/%Y"
            )
        if "orig_index" in df_out.columns:
            df_out.drop(columns=["orig_index"], inplace=True)

        df_out.to_excel(writer, sheet_name="Original", index=False)
        ws = writer.sheets["Original"]

        # --- Color the Debit/Credit/Difference cells by match group ---
        col_map = {name: idx + 1 for idx, name in enumerate(df_out.columns)}
        debit_col = col_map.get("Debit")
        credit_col = col_map.get("Credit")
        diff_col = col_map.get("Difference")
        group_col = col_map.get("Gruppo")

        for excel_row, m in enumerate(meta, start=2):
            if m.get("month_total"):
                self._style_month_total_row(
                    ws, excel_row, df_out.columns, m
                )
                continue
            if m["color_idx"] is None:
                continue

            light = PatternFill(
                start_color=GROUP_FILLS[m["color_idx"]][0],
                end_color=GROUP_FILLS[m["color_idx"]][0],
                fill_type="solid",
            )
            if m["side"] == "debit" and debit_col:
                ws.cell(row=excel_row, column=debit_col).fill = light
            elif m["side"] == "credit" and credit_col:
                ws.cell(row=excel_row, column=credit_col).fill = light
            if diff_col:
                cell = ws.cell(row=excel_row, column=diff_col)
                cell.fill = PatternFill(
                    start_color=GROUP_FILLS[m["color_idx"]][1],
                    end_color=GROUP_FILLS[m["color_idx"]][1],
                    fill_type="solid",
                )
                cell.number_format = "#,##0.00 €"
                cell.font = Font(
                    bold=True,
                    color="C00000" if (m["difference"] or 0) > 0 else "000000",
                )
            if group_col:
                font = (
                    Font(bold=True, italic=True, color="1F4E78")
                    if m["inserted"]
                    else Font(bold=True)
                )
                ws.cell(row=excel_row, column=group_col).font = font

        self._add_original_legend(ws, len(df_out), saldo_analysis)

    @staticmethod
    def _style_month_total_row(ws, excel_row, columns, m):
        """Applies the subtotal style to a 'TOTALE MESE' row: bold text,
        light gray fill across the row and currency format on the amounts."""
        col_map = {name: idx + 1 for idx, name in enumerate(columns)}
        fill = PatternFill(
            start_color="E7E6E6", end_color="E7E6E6", fill_type="solid"
        )
        font = Font(bold=True)
        for name in columns:
            cell = ws.cell(row=excel_row, column=col_map[name])
            cell.fill = fill
            cell.font = font
            if name in ("Debit", "Credit", "Difference"):
                cell.number_format = "#,##0.00 €"
        diff_cell = ws.cell(row=excel_row, column=col_map.get("Difference", 1))
        diff_cell.font = Font(bold=True, color="C00000" if m["difference"] != 0 else "000000")

    @staticmethod
    def _row_meta(member, check, inserted):
        """Returns the per-row metadata used to color/style the Original sheet."""
        if member is None:
            return {
                "color_idx": None,
                "side": None,
                "difference": None,
                "group_label": "",
                "check": check,
                "inserted": inserted,
            }
        return {
            "color_idx": member["color_idx"],
            "side": member["side"],
            "difference": member["difference"],
            "group_label": member["transaction_id"],
            "check": check,
            "inserted": inserted,
        }

    def _create_monthly_balance_sheet(self, writer):
        df = self.engine._calculate_monthly_balance()
        if df.empty:
            return

        for c in df.columns:
            if c != "Month":
                df[c] = df[c] / 100.0

        df.to_excel(writer, sheet_name="Monthly Balance", index=False)
        ws = writer.sheets["Monthly Balance"]

        # Rename headers to Italian for operator clarity
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_align = Alignment(horizontal="center")
        italian_headers = {
            "Month": "Mese",
            "Total Debit": "Incassi Totali",
            "Used Debit": "Incassi Riconc.",
            "Total Credit": "Versamenti Totali",
            "Used Credit": "Versamenti Riconc.",
            "Differenza Mensile": "Differenza Mensile",
            "Cumulato": "Cumulato",
            "Versamenti Non Agganciati": "Vers. Non Agganciati",
        }
        for col_idx, cell in enumerate(ws[1], start=1):
            header = cell.value
            if header in italian_headers:
                cell.value = italian_headers[header]
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align

        # Apply currency style to amount columns (B-H)
        for row in ws.iter_rows(min_row=2, max_row=len(df) + 1, min_col=2, max_col=8):
            for cell in row:
                cell.style = self.currency_style

        # Conditional formatting: red background for "Vers. Non Agganciati" > 0
        red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
        non_agg_range = f"H2:H{len(df) + 1}"
        ws.conditional_formatting.add(
            non_agg_range,
            CellIsRule(operator="greaterThan", formula=["0"], fill=red_fill),
        )

        # Cumulato: conditional green if >= 0, red if < 0
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        red_fill_light = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        cumulato_range = f"G2:G{len(df) + 1}"
        ws.conditional_formatting.add(
            cumulato_range,
            CellIsRule(operator="greaterThanOrEqual", formula=["0"], fill=green_fill),
        )
        ws.conditional_formatting.add(
            cumulato_range,
            CellIsRule(operator="lessThan", formula=["0"], fill=red_fill_light),
        )

        # Set column widths
        ws.column_dimensions["A"].width = 12
        for col_letter in ["B", "C", "D", "E", "F", "G", "H"]:
            ws.column_dimensions[col_letter].width = 20
