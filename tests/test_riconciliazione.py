import unittest
import os
import sys
import pandas as pd
from datetime import datetime

# Add root folder to sys.path
sys.path.append(os.path.join(os.path.dirname(__file__), '..'))

from core import ReconciliationEngine


class TestReconciliationCore(unittest.TestCase):
    """
    Unit test suite for ReconciliationEngine in core.py.
    """

    def _create_df(self, debits_list, credits_list):
        """Helper to create input DataFrame in the format expected by core.py."""
        rows = []
        for d, amt in debits_list:
            rows.append({'Date': d, 'Debit': amt, 'Credit': 0.0})
        for d, amt in credits_list:
            rows.append({'Date': d, 'Debit': 0.0, 'Credit': amt})
        
        df = pd.DataFrame(rows)
        return df

    def test_exact_1_to_1_match(self):
        """Verifies a simple 1-to-1 match."""
        credits = [(datetime(2025, 1, 10), 100.0)]
        debits = [
            (datetime(2025, 1, 9), 50.0),
            (datetime(2025, 1, 9), 100.0),
            (datetime(2025, 1, 11), 20.0)
        ]

        df = self._create_df(debits, credits)
        engine = ReconciliationEngine(
            tolerance=0.0,
            days_window=5,
            search_direction="past_only",
            algorithm="subset_sum"
        )
        engine.run(df, verbose=False)

        matches = engine.matches_df
        self.assertEqual(len(matches), 1, "There should be exactly one match.")
        self.assertEqual(matches.iloc[0]['total_credit'], 10000)  # Cents

        # Check used flags
        self.assertTrue(engine.debit_df.loc[engine.debit_df['orig_index'] == 1, 'used'].values[0])
        self.assertTrue(engine.credit_df.loc[engine.credit_df['orig_index'] == 3, 'used'].values[0])

    def test_nearest_date_match(self):
        """Verifies that among multiple exact matches, the candidate closest in date is selected."""
        debits = [(datetime(2025, 1, 1), 100.0)]
        credits = [
            (datetime(2025, 1, 10), 100.0),  # 9 days diff
            (datetime(2025, 1, 2), 100.0),   # 1 day diff (closest)
        ]

        df = self._create_df(debits, credits)
        engine = ReconciliationEngine(
            tolerance=0.0,
            days_window=10,
            search_direction="past_only",
            algorithm="subset_sum"
        )
        engine.run(df, verbose=False)

        matches = engine.matches_df
        self.assertEqual(len(matches), 1)
        self.assertEqual(matches.iloc[0]['credit_indices'], [2])

    def test_2_to_1_combination(self):
        """Verifies a combined 2 DEBITs to 1 CREDIT match."""
        credits = [(datetime(2025, 2, 15), 150.0)]
        debits = [
            (datetime(2025, 2, 14), 100.0),
            (datetime(2025, 2, 14), 50.0),
            (datetime(2025, 2, 16), 150.0)
        ]

        df = self._create_df(debits, credits)
        engine = ReconciliationEngine(
            max_combinations=2,
            tolerance=0.0,
            days_window=5,
            search_direction="past_only",
            algorithm="subset_sum"
        )
        engine.run(df, verbose=False)

        matches = engine.matches_df
        self.assertEqual(len(matches), 1)
        self.assertEqual(len(matches.iloc[0]['debit_indices']), 2)
        self.assertEqual(matches.iloc[0]['total_credit'], 15000)

    def test_match_with_tolerance(self):
        """Verifies that a match occurs within the defined tolerance.

        With the residual logic, a 99.99 deposit covered by a 100.00 receipt
        consumes only 99.99 of the receipt: the match is exact (difference 0)
        and the 0.01 residual stays available.
        """
        credits = [(datetime(2025, 3, 10), 99.99)]
        debits = [(datetime(2025, 3, 10), 100.00)]

        df = self._create_df(debits, credits)
        engine = ReconciliationEngine(tolerance=0.02)
        engine.run(df, verbose=False)

        matches = engine.matches_df
        self.assertEqual(len(matches), 1)
        self.assertEqual(matches.iloc[0]['total_debit'], 9999)  # used amount only
        self.assertEqual(matches.iloc[0]['difference'], 0)

    def test_no_match_found(self):
        """Verifies that no match is created if there are no valid candidates."""
        credits = [(datetime(2025, 4, 1), 1000.0)]
        debits = [(datetime(2025, 4, 1), 100.0)]

        df = self._create_df(debits, credits)
        engine = ReconciliationEngine(
            tolerance=0.0,
            enable_best_fit=False,
            algorithm="subset_sum"
        )
        engine.run(df, verbose=False)

        matches = engine.matches_df
        self.assertTrue(matches.empty)
        self.assertFalse(engine.credit_df['used'].any())

    def test_progressive_balance_operator_defaults(self):
        """Verifies Progressive Balance with POS operator default settings."""
        credits = [(datetime(2025, 5, 5), 300.0)]
        debits = [
            (datetime(2025, 5, 2), 100.0),
            (datetime(2025, 5, 3), 200.0)
        ]

        df = self._create_df(debits, credits)
        engine = ReconciliationEngine(
            algorithm="progressive_balance",
            days_window=5,
            tolerance=50.0,
            search_direction="past_only"
        )
        engine.run(df, verbose=False)

        matches = engine.matches_df
        self.assertEqual(len(matches), 1)
        self.assertEqual(matches.iloc[0]['total_debit'], 30000)
        self.assertEqual(matches.iloc[0]['total_credit'], 30000)

    def test_time_window_compliance(self):
        """Verifies that receipts outside days_window are not matched."""
        credits = [(datetime(2025, 7, 15), 200.0)]
        debits = [(datetime(2025, 7, 1), 200.0)]  # 14 days prior

        df = self._create_df(debits, credits)
        engine = ReconciliationEngine(
            tolerance=0.0,
            days_window=5,
            residual_days_window=5,
            search_direction="past_only",
            algorithm="subset_sum"
        )
        engine.run(df, verbose=False)

        matches = engine.matches_df
        self.assertTrue(matches.empty, "No match should be found outside window.")

    def test_max_combinations_limit(self):
        """Verifies combination limits are respected."""
        credits = [(datetime(2025, 8, 10), 60.0)]
        debits = [
            (datetime(2025, 8, 9), 10.0),
            (datetime(2025, 8, 9), 20.0),
            (datetime(2025, 8, 9), 30.0)
        ]

        df = self._create_df(debits, credits)
        engine = ReconciliationEngine(
            tolerance=0.0,
            max_combinations=2,
            enable_best_fit=False,
            algorithm="subset_sum"
        )
        engine.run(df, verbose=False)

        matches = engine.matches_df
        self.assertTrue(matches.empty)


class TestReportingVisualization(unittest.TestCase):
    """
    Unit tests for the visualization/quadratura features in reporting.py
    (color grouping in the Original sheet and Saldo Prog. analysis).
    """

    @classmethod
    def setUpClass(cls):
        from reporting import ExcelReporter

        cls.reporter_cls = ExcelReporter

    def _run_engine(self):
        """Creates a small engine with two match groups."""
        rows = [
            {"Date": datetime(2025, 1, 2), "Debit": 100.0, "Credit": 0.0},
            {"Date": datetime(2025, 1, 2), "Debit": 200.0, "Credit": 0.0},
            {"Date": datetime(2025, 1, 3), "Debit": 0.0, "Credit": 150.0},
            {"Date": datetime(2025, 1, 4), "Debit": 300.0, "Credit": 0.0},
            {"Date": datetime(2025, 1, 5), "Debit": 0.0, "Credit": 200.0},
        ]
        df = pd.DataFrame(rows)
        engine = ReconciliationEngine(
            tolerance=10.0,
            days_window=5,
            search_direction="past_only",
            algorithm="progressive_balance",
        )
        engine.run(df, verbose=False)
        self.assertGreaterEqual(len(engine.matches_df), 1)
        return engine

    def test_build_match_groups_color_cycle(self):
        """Colors repeat every 3 groups: groups 1,4,7 share color index 0."""
        engine = self._run_engine()
        reporter = self.reporter_cls(engine)
        groups = reporter._build_match_groups()
        self.assertTrue(groups, "At least one group should be built.")

        non_split = [i for i in groups.values() if not i.get("split")]
        self.assertTrue(non_split, "At least one non-split row expected.")
        for info in non_split:
            self.assertEqual(
                info["color_idx"], (info["group_id"] - 1) % 3,
                "Color must cycle every 3 groups.",
            )

    def test_split_rows_keep_first_group_color(self):
        """Rows shared by multiple groups keep the FIRST group's color and list
        the additional Transaction IDs after a ✂ marker."""
        engine = self._run_engine()
        reporter = self.reporter_cls(engine)
        groups = reporter._build_match_groups()

        split = [i for i in groups.values() if i.get("split")]
        if split:
            for info in split:
                self.assertIsNotNone(info["color_idx"])
                self.assertGreater(len(info["other_transaction_ids"]), 0)
                # each membership records the amount consumed by its group
                self.assertTrue(
                    all(m.get("amount") is not None for m in info["memberships"])
                )

    def test_original_sheet_splits_residual_rows(self):
        """A receipt split across deposits is shown on the original row (first
        group's portion) plus one inserted row per residual portion, with the
        same date and its own group's color. Total amounts are preserved."""
        rows = [
            {"Date": datetime(2025, 1, 2), "Debit": 2535.50, "Credit": 0.0},
            {"Date": datetime(2025, 1, 3), "Debit": 2777.00, "Credit": 0.0},
            {"Date": datetime(2025, 1, 4), "Debit": 1100.00, "Credit": 0.0},
            {"Date": datetime(2025, 1, 5), "Debit": 4015.00, "Credit": 0.0},
            {"Date": datetime(2025, 1, 5), "Debit": 0.0, "Credit": 3000.00},
            {"Date": datetime(2025, 1, 7), "Debit": 0.0, "Credit": 7000.00},
        ]
        df = pd.DataFrame(rows)
        engine = ReconciliationEngine(
            tolerance=50.0,
            days_window=5,
            search_direction="past_only",
            algorithm="progressive_balance",
        )
        out = "/tmp/opencode/test_split_report.xlsx"
        engine.run(df, output_file=out, verbose=False)

        # find a split receipt: same orig_index in 2+ groups
        groups = self.reporter_cls(engine)._build_match_groups()
        split_ois = [
            oi for oi, i in groups.items()
            if i.get("side") == "debit" and len(i["memberships"]) > 1
        ]
        self.assertTrue(split_ois, "Expected at least one split receipt.")
        for oi in split_ois:
            info = groups[oi]
            tids = [m["transaction_id"] for m in info["memberships"]]
            self.assertEqual(len(set(tids)), len(tids), "Each portion has its own group.")
            self.assertEqual(info["transaction_id"], tids[0], "Original row keeps first group.")

        # check amounts are preserved in the Original sheet
        from openpyxl import load_workbook

        wb = load_workbook(out)
        ws = wb["Original"]
        total_debit = sum(
            v
            for r in range(2, ws.max_row + 1)
            if isinstance(v := ws.cell(row=r, column=2).value, (int, float))
        )
        self.assertAlmostEqual(total_debit, 10427.50, places=2)
        self.assertGreater(ws.max_row - 1, len(df), "Inserted residual rows expected.")

    def test_group_members_share_transaction_and_difference(self):
        """All members of a single group carry the same Transaction ID and delta."""
        engine = self._run_engine()
        reporter = self.reporter_cls(engine)
        groups = reporter._build_match_groups()

        by_group = {}
        for info in groups.values():
            by_group.setdefault(info["group_id"], []).append(info)
        for group_id, members in by_group.items():
            ids = {m["transaction_id"] for m in members}
            diffs = {m["difference"] for m in members}
            self.assertEqual(len(ids), 1, "Group members must share the same Transaction ID.")
            self.assertEqual(len(diffs), 1, "Group members must share the same difference.")

    def test_saldo_prog_analysis_present(self):
        """Saldo Prog. analysis detects a consistent running cash balance.
        Debit/Credit are in cents, Saldo Prog. is in euros (as in the pipeline)."""
        df = pd.DataFrame(
            [
                {"Date": datetime(2025, 1, 2), "Debit": 0, "Credit": 10000,
                 "Saldo Prog.": 900.0, "orig_index": 0},
                {"Date": datetime(2025, 1, 2), "Debit": 20000, "Credit": 0,
                 "Saldo Prog.": 1100.0, "orig_index": 1},
                {"Date": datetime(2025, 1, 3), "Debit": 0, "Credit": 15000,
                 "Saldo Prog.": 950.0, "orig_index": 2},
            ]
        )
        reporter = self.reporter_cls(None)
        analysis = reporter._analyze_saldo_prog(df)

        self.assertTrue(analysis["present"])
        self.assertEqual(analysis["opening"], 1000.0)  # 900 - (0 - 100)
        self.assertEqual(analysis["closing"], 950.0)
        self.assertEqual(analysis["inconsistent_rows"], 0)
        self.assertEqual(analysis["negative_rows"], 0)

    def test_saldo_prog_analysis_missing_column(self):
        """Saldo Prog. analysis is skipped when the column is absent."""
        df = pd.DataFrame(
            [
                {"Date": datetime(2025, 1, 2), "Debit": 0, "Credit": 10000,
                 "orig_index": 0},
            ]
        )
        reporter = self.reporter_cls(None)
        analysis = reporter._analyze_saldo_prog(df)
        self.assertFalse(analysis["present"])

    def test_saldo_prog_analysis_inconsistent(self):
        """Saldo Prog. analysis flags rows that do not match the theoretical cash."""
        df = pd.DataFrame(
            [
                {"Date": datetime(2025, 1, 2), "Debit": 0, "Credit": 10000,
                 "Saldo Prog.": 900.0, "orig_index": 0},
                {"Date": datetime(2025, 1, 2), "Debit": 20000, "Credit": 0,
                 "Saldo Prog.": 999.0, "orig_index": 1},  # should be 1100
            ]
        )
        reporter = self.reporter_cls(None)
        analysis = reporter._analyze_saldo_prog(df)

        self.assertTrue(analysis["present"])
        self.assertEqual(analysis["inconsistent_rows"], 1)
        self.assertIn("Saldo incoerente", analysis["check_map"][1])


if __name__ == '__main__':
    unittest.main(verbosity=2)
